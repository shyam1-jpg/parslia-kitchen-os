#!/usr/bin/env python3
"""Tests for Sham/Shyam overtime totals and 1500 / 1800 flags."""

from __future__ import annotations

import csv
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

from overtime_hours import (
    PERIOD_END,
    PERIOD_START,
    HourRow,
    contracted_hours,
    extra_hours,
    flags_for,
    is_shyam,
    parse_csv_file,
    parse_date,
    parse_hours,
    parse_xlsx_file,
    summarise,
    verdict_text,
    write_workbook,
    ScanResult,
)


class NameMatching(unittest.TestCase):
    def test_shyam_and_sham(self):
        self.assertTrue(is_shyam("Shyam Prasad"))
        self.assertTrue(is_shyam("sham"))
        self.assertTrue(is_shyam("SHAM PRASAD"))
        self.assertTrue(is_shyam("S. Prasad"))

    def test_not_shannon_or_other_staff(self):
        self.assertFalse(is_shyam("Shannon"))
        self.assertFalse(is_shyam("Saurabh Unial"))
        self.assertFalse(is_shyam(""))
        self.assertFalse(is_shyam(None))


class Parsers(unittest.TestCase):
    def test_hours_formats(self):
        self.assertEqual(parse_hours(8), 8.0)
        self.assertEqual(parse_hours("8.5"), 8.5)
        self.assertEqual(parse_hours("8:30"), 8.5)
        self.assertEqual(parse_hours("8h 15m"), 8.25)
        self.assertIsNone(parse_hours(""))

    def test_dates(self):
        self.assertEqual(parse_date("2024-03-01"), date(2024, 3, 1))
        self.assertEqual(parse_date("01/03/2024"), date(2024, 3, 1))
        self.assertEqual(parse_date(date(2023, 6, 2)), date(2023, 6, 2))


class Totals(unittest.TestCase):
    def test_contracted_weekdays_only(self):
        # Mon 2 Jan 2023 – Fri 6 Jan 2023 = 5 weekdays
        self.assertEqual(contracted_hours(date(2023, 1, 2), date(2023, 1, 6)), 40.0)
        # Includes Saturday
        self.assertEqual(contracted_hours(date(2023, 1, 2), date(2023, 1, 7)), 40.0)

    def test_below_1500(self):
        extra = extra_hours(8000, 7600)
        self.assertEqual(extra, 400)
        self.assertFalse(flags_for(extra)["above_1500"])
        self.assertFalse(flags_for(extra)["above_1800"])

    def test_between_1500_and_1800(self):
        extra = extra_hours(9200, 7600)
        self.assertEqual(extra, 1600)
        self.assertTrue(flags_for(extra)["above_1500"])
        self.assertFalse(flags_for(extra)["above_1800"])

    def test_above_1800(self):
        extra = extra_hours(9600, 7600)
        self.assertEqual(extra, 2000)
        self.assertTrue(flags_for(extra)["above_1500"])
        self.assertTrue(flags_for(extra)["above_1800"])

    def test_summarise_filters_years_and_shannon(self):
        rows = [
            HourRow("f", "Shyam Prasad", date(2022, 12, 30), 10, 2),
            HourRow("f", "Shyam Prasad", date(2024, 6, 3), 12, 4),
            HourRow("f", "Shyam Prasad", date(2026, 8, 26), 10, 2),
            HourRow("f", "Shyam Prasad", date(2026, 8, 27), 10, 2),
        ]
        summary = summarise(rows)
        self.assertEqual(summary["dated_rows"], 2)
        self.assertEqual(summary["actual_hours"], 22)
        self.assertEqual(summary["declared_overtime"], 6)
        self.assertEqual(summary["by_year"][2024]["actual_hours"], 12)
        self.assertEqual(summary["by_year"][2023]["rows"], 0)

    def test_cannot_confirm_without_rows(self):
        status, detail = verdict_text(summarise([]))
        self.assertEqual(status, "CANNOT CONFIRM")
        self.assertIn("1500", detail)


class FileImport(unittest.TestCase):
    def test_csv_cypad_style(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "cypad-timesheet.csv"
            with path.open("w", newline="", encoding="utf-8") as fh:
                w = csv.writer(fh)
                w.writerow(["Staff Name", "Date", "Hours", "Overtime Hours"])
                w.writerow(["Shannon", "03/01/2024", "12", "4"])
                w.writerow(["Shyam Prasad", "03/01/2024", "11", "3"])
                w.writerow(["Sham Prasad", "04/01/2024", "10", "2"])
            rows = parse_csv_file(path)
            self.assertEqual(len(rows), 2)
            self.assertEqual(sum(r.hours for r in rows), 21)

    def test_xlsx_clock_in_out(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "hours.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Employee", "Date", "Clock In", "Clock Out"])
            ws.append(["Shyam Prasad", "2025-02-03", "08:00", "18:00"])
            ws.append(["Chetan", "2025-02-03", "08:00", "20:00"])
            wb.save(path)
            rows = parse_xlsx_file(path)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0].hours, 10.0)

    def test_workbook_verdict_sheet(self):
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "out.xlsx"
            scan = ScanResult()
            summary = summarise([])
            write_workbook(out, scan, summary, ["/workspace"])
            wb = load_workbook(out, data_only=True)
            self.assertIn("Verdict", wb.sheetnames)
            self.assertIn("Year Totals 2023-2026", wb.sheetnames)
            self.assertIn("Paste Cypad Timesheet", wb.sheetnames)
            verdict = wb["Verdict"]
            self.assertEqual(verdict["B16"].value, "CANNOT CONFIRM")
            years = wb["Year Totals 2023-2026"]
            self.assertEqual(years["A2"].value, 2023)
            self.assertEqual(years["A5"].value, 2026)
            self.assertLessEqual(years["C5"].value.date(), PERIOD_END)
            self.assertGreaterEqual(years["B2"].value.date(), PERIOD_START)


if __name__ == "__main__":
    unittest.main()
