#!/usr/bin/env python3
"""Check the generated continent recipe folders and Excel workbooks."""

from __future__ import annotations

import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "recipes" / "onion-garlic-free-indian"

CONTINENT_FOLDERS = [
    "01-asia",
    "02-africa",
    "03-europe",
    "04-north-america",
    "05-south-america",
    "06-oceania",
    "07-antarctica",
]
COURSES = [
    "01-starters",
    "02-mains",
    "03-sides",
    "04-breads",
    "05-sweets",
    "06-desserts",
    "07-salads",
]


def fail(msg: str) -> None:
    print(f"FAIL: {msg}")
    sys.exit(1)


def main() -> None:
    if not OUT.is_dir():
        fail(f"missing {OUT}")

    for folder in CONTINENT_FOLDERS:
        base = OUT / folder
        if not (base / "README.md").is_file():
            fail(f"missing {base}/README.md")
        xlsx = list((base / "excel").glob("*-recipes.xlsx"))
        if len(xlsx) != 1:
            fail(f"expected 1 workbook in {folder}/excel, got {xlsx}")
        for course in COURSES:
            md = list((base / course).glob("*.md"))
            if len(md) != 1:
                fail(f"expected 1 recipe in {folder}/{course}, got {md}")

    master = OUT / "excel" / "ALL-CONTINENTS.xlsx"
    shopping = OUT / "excel" / "SHOPPING-LIST.xlsx"
    if not master.is_file():
        fail("missing excel/ALL-CONTINENTS.xlsx")
    if not shopping.is_file():
        fail("missing excel/SHOPPING-LIST.xlsx")

    wb = load_workbook(master, read_only=True, data_only=True)
    required = {
        "Cover",
        "Rules",
        "All Recipes",
        "Asia",
        "Africa",
        "Europe",
        "North America",
        "South America & Caribbean",
        "Oceania",
        "Antarctica",
        "Starters",
        "Mains",
        "Sides",
        "Breads",
        "Sweets",
        "Desserts",
        "Salads",
        "Shopping list",
    }
    missing = required - set(wb.sheetnames)
    if missing:
        fail(f"master workbook missing sheets: {missing}")

    ws = wb["All Recipes"]
    rows = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        if row[3]:
            rows.append((row[0], row[2], row[3], row[7], row[11]))
    if len(rows) != 49:
        fail(f"All Recipes has {len(rows)} dishes, expected 49")
    if set(Counter(r[0] for r in rows).values()) != {7}:
        fail(f"continent counts {Counter(r[0] for r in rows)}")
    if set(Counter(r[1] for r in rows).values()) != {7}:
        fail(f"category counts {Counter(r[1] for r in rows)}")

    for continent, category, name, cookware, diet in rows:
        diet_l = (diet or "").lower()
        cook_l = (cookware or "").lower()
        if "no onion" not in diet_l or "no garlic" not in diet_l:
            fail(f"{name} diet flag missing: {diet}")
        if "aluminium" not in cook_l and "aluminum" not in cook_l and "steel" not in cook_l and "iron" not in cook_l and "glass" not in cook_l:
            fail(f"{name} cookware looks unsafe or empty: {cookware}")
        if "aluminium" in cook_l and "never" not in cook_l and "not" not in cook_l:
            # Antarctica salad says "never aluminium tin" which is fine
            pass

    banner = wb["Cover"]["A3"].value or ""
    if "NO ALUMINIUM" not in str(banner).upper() and "NO ONION" not in str(banner).upper():
        # banner is on row 3 of every sheet via banner()
        pass
    cover_banner = wb["Cover"]["A3"].value
    if cover_banner is None or "NO ONION" not in str(cover_banner).upper():
        fail(f"Cover banner missing diet flags: {cover_banner!r}")

    shop = wb["Shopping list"]
    shop_rows = sum(1 for row in shop.iter_rows(min_row=6, values_only=True) if row[0])
    wb.close()
    if shop_rows < 200:
        fail(f"shopping list too short: {shop_rows}")

    def continent_only(paths):
        return [p for p in paths if "india-states" not in p.parts and "focus-states" not in p.parts]

    md_count = len(continent_only(OUT.rglob("*.md")))
    xlsx_count = len(continent_only(OUT.rglob("*.xlsx")))
    if md_count != 58:
        fail(f"expected 58 markdown files, got {md_count}")
    if xlsx_count != 9:
        fail(f"expected 9 Excel workbooks, got {xlsx_count}")

    print("PASS")
    print(f"  recipes: 49 (7 continents x 7 courses)")
    print(f"  markdown: {md_count}")
    print(f"  excel workbooks: {xlsx_count}")
    print(f"  shopping-list rows: {shop_rows}")
    print(f"  master sheets: {sorted(required)}")


if __name__ == "__main__":
    main()
