#!/usr/bin/env python3
"""Find and total Sham/Shyam overtime hours from Excel/CSV (2023–2026).

Cloud agents cannot read a Windows C: or D: drive. This script:
  1. Scans this environment for .xlsx / .xls / .csv timesheets
  2. Totals hours for Shyam Prasad (not Shannon)
  3. Compares extra hours against 1500 and 1800
  4. Writes D_Drive/Shyam_Overtime_Hours_2023-2026.xlsx
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable

PERIOD_START = date(2023, 1, 1)
PERIOD_END = date(2026, 8, 26)
CONTRACTED_HOURS_PER_DAY = 8.0
THRESHOLDS = (1500.0, 1800.0)
STAFF_NAME = "Shyam Prasad"
STAFF_ROLE = "Head Chef"

# Match Sham / Shyam / S. Prasad, but never Shannon.
NAME_RE = re.compile(
    r"\b(?:shyam|sham)\b|\bs\.?\s*prasad\b|\bprasad\b.*\b(?:shyam|sham)\b",
    re.I,
)
SHANNON_RE = re.compile(r"\bshannon\b", re.I)

NAME_HEADERS = {"name", "employee", "staff", "person", "worker", "user", "full name", "employee name", "staff name"}
DATE_HEADERS = {"date", "day", "shift date", "work date", "clock date"}
HOURS_HEADERS = {
    "hours",
    "hrs",
    "total hours",
    "total hrs",
    "worked",
    "worked hours",
    "duration",
    "time worked",
}
OT_HEADERS = {"overtime", "ot", "ot hours", "overtime hours", "extra hours", "extra"}
IN_HEADERS = {"clock in", "in", "start", "start time", "time in"}
OUT_HEADERS = {"clock out", "out", "end", "end time", "time out"}

SKIP_DIR_NAMES = {
    ".git",
    "node_modules",
    ".npm",
    ".cache",
    ".nvm",
    "DerivedData",
    ".cursor-server",
    ".local",
    "pkg",
    "testdata",
}


@dataclass
class HourRow:
    source: str
    person: str
    work_date: date | None
    hours: float
    overtime_hours: float | None
    notes: str = ""


@dataclass
class ScanResult:
    files_seen: list[str] = field(default_factory=list)
    timesheet_files: list[str] = field(default_factory=list)
    rows: list[HourRow] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def is_shyam(name: str | None) -> bool:
    if not name:
        return False
    text = str(name).strip()
    if not text or SHANNON_RE.search(text):
        return False
    return bool(NAME_RE.search(text))


def parse_hours(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip().lower().replace(",", "")
    if not text or text in {"n/a", "na", "-", "—"}:
        return None
    hm = re.fullmatch(r"(\d{1,3}):([0-5]\d)(?::[0-5]\d)?", text)
    if hm:
        return int(hm.group(1)) + int(hm.group(2)) / 60.0
    labeled = re.fullmatch(r"(?:(\d+)\s*h(?:ours?)?)?\s*(?:(\d+)\s*m(?:ins?)?)?", text)
    if labeled and (labeled.group(1) or labeled.group(2)):
        hours = int(labeled.group(1) or 0)
        mins = int(labeled.group(2) or 0)
        return hours + mins / 60.0
    try:
        return float(text)
    except ValueError:
        return None


def parse_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and 20000 < float(value) < 60000:
        # Excel serial date
        return date(1899, 12, 30) + timedelta(days=int(value))
    text = str(value).strip()
    for fmt in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%d/%m/%y",
        "%Y/%m/%d",
        "%d %b %Y",
        "%d %B %Y",
        "%m/%d/%Y",
    ):
        try:
            return datetime.strptime(text[:32], fmt).date()
        except ValueError:
            continue
    return None


def in_period(d: date | None) -> bool:
    if d is None:
        return False
    return PERIOD_START <= d <= PERIOD_END


def weekdays_in_range(start: date, end: date) -> int:
    if end < start:
        return 0
    n = 0
    d = start
    while d <= end:
        if d.weekday() < 5:
            n += 1
        d += timedelta(days=1)
    return n


def contracted_hours(start: date, end: date) -> float:
    return weekdays_in_range(start, end) * CONTRACTED_HOURS_PER_DAY


def extra_hours(actual: float, contracted: float) -> float:
    return round(actual - contracted, 2)


def flags_for(extra: float) -> dict[str, bool]:
    return {f"above_{int(t)}": extra >= t for t in THRESHOLDS}


def year_windows() -> list[tuple[int, date, date]]:
    windows = []
    for year in (2023, 2024, 2025, 2026):
        start = date(year, 1, 1)
        end = date(year, 12, 31)
        if start < PERIOD_START:
            start = PERIOD_START
        if end > PERIOD_END:
            end = PERIOD_END
        if start <= end:
            windows.append((year, start, end))
    return windows


def _norm_header(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _header_map(headers: Iterable) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for i, raw in enumerate(headers):
        key = _norm_header(raw)
        if key:
            mapping[key] = i
    return mapping


def _pick(mapping: dict[str, int], names: set[str]) -> int | None:
    for name in names:
        if name in mapping:
            return mapping[name]
    return None


def _hours_from_clock(clock_in, clock_out) -> float | None:
    t1 = parse_hours(clock_in)
    t2 = parse_hours(clock_out)
    # Clock times like 09:00 are 9.0 hours-from-midnight, not durations.
    if t1 is None or t2 is None:
        return None
    if t2 >= t1:
        return round(t2 - t1, 2)
    return round((24 - t1) + t2, 2)


def parse_tabular_rows(headers, records, source: str) -> list[HourRow]:
    mapping = _header_map(headers)
    name_i = _pick(mapping, NAME_HEADERS)
    date_i = _pick(mapping, DATE_HEADERS)
    hours_i = _pick(mapping, HOURS_HEADERS)
    ot_i = _pick(mapping, OT_HEADERS)
    in_i = _pick(mapping, IN_HEADERS)
    out_i = _pick(mapping, OUT_HEADERS)
    if name_i is None:
        return []
    if hours_i is None and ot_i is None and not (in_i is not None and out_i is not None):
        return []

    rows: list[HourRow] = []
    for rec in records:
        rec = list(rec)
        if name_i >= len(rec):
            continue
        person = rec[name_i]
        if not is_shyam(person):
            continue
        work_date = parse_date(rec[date_i]) if date_i is not None and date_i < len(rec) else None
        hours = parse_hours(rec[hours_i]) if hours_i is not None and hours_i < len(rec) else None
        ot = parse_hours(rec[ot_i]) if ot_i is not None and ot_i < len(rec) else None
        if hours is None and in_i is not None and out_i is not None:
            hours = _hours_from_clock(
                rec[in_i] if in_i < len(rec) else None,
                rec[out_i] if out_i < len(rec) else None,
            )
        if hours is None and ot is None:
            continue
        notes = " ".join(str(x) for x in rec if x is not None)
        if "EXAMPLE ROW" in notes.upper():
            continue
        rows.append(
            HourRow(
                source=source,
                person=str(person).strip(),
                work_date=work_date,
                hours=float(hours or 0.0),
                overtime_hours=ot,
                notes="imported",
            )
        )
    return rows


def parse_csv_file(path: Path) -> list[HourRow]:
    raw = path.read_bytes()
    text = None
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        return []
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows = list(reader)
    if not rows:
        return []
    return parse_tabular_rows(rows[0], rows[1:], str(path))


def parse_xlsx_file(path: Path) -> list[HourRow]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    found: list[HourRow] = []
    try:
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            found.extend(parse_tabular_rows(rows[0], rows[1:], f"{path}#{ws.title}"))
    finally:
        wb.close()
    return found


def looks_like_timesheet(path: Path) -> bool:
    name = path.name.lower()
    return any(token in name for token in ("hour", "time", "clock", "cypad", "rota", "overtime", "shift", "staff"))


def scan_files(roots: Iterable[Path], extra_files: Iterable[Path] = ()) -> ScanResult:
    result = ScanResult()
    seen: set[Path] = set()

    def consider(path: Path) -> None:
        try:
            path = path.resolve()
        except OSError:
            return
        if path in seen or not path.is_file():
            return
        seen.add(path)
        suffix = path.suffix.lower()
        if suffix not in {".xlsx", ".xls", ".csv"}:
            return
        if path.name.lower().startswith("shyam_overtime_hours"):
            return
        result.files_seen.append(str(path))
        try:
            rows = parse_xlsx_file(path) if suffix == ".xlsx" else parse_csv_file(path) if suffix == ".csv" else []
        except Exception as exc:  # pragma: no cover - defensive
            result.errors.append(f"{path}: {exc}")
            return
        if rows or looks_like_timesheet(path):
            result.timesheet_files.append(str(path))
        result.rows.extend(rows)

    for extra in extra_files:
        consider(Path(extra))

    for root in roots:
        root = Path(root)
        if not root.exists():
            continue
        if root.is_file():
            consider(root)
            continue
        for path in root.rglob("*"):
            if any(part in SKIP_DIR_NAMES for part in path.parts):
                continue
            consider(path)
    return result


def summarise(rows: list[HourRow]) -> dict:
    in_range = [r for r in rows if in_period(r.work_date)]
    undated = [r for r in rows if r.work_date is None]
    by_year: dict[int, dict] = {}
    for year, start, end in year_windows():
        year_rows = [r for r in in_range if r.work_date and start <= r.work_date <= end]
        actual = round(sum(r.hours for r in year_rows), 2)
        ot_declared = [r.overtime_hours for r in year_rows if r.overtime_hours is not None]
        contracted = contracted_hours(start, end)
        extra = extra_hours(actual, contracted) if year_rows else None
        declared_ot = round(sum(ot_declared), 2) if ot_declared else None
        by_year[year] = {
            "start": start,
            "end": end,
            "rows": len(year_rows),
            "actual_hours": actual,
            "contracted_hours": round(contracted, 2),
            "extra_hours": extra,
            "declared_overtime": declared_ot,
            "flags": flags_for(extra if extra is not None else 0) if extra is not None else None,
        }

    actual_total = round(sum(r.hours for r in in_range), 2)
    contracted_total = round(contracted_hours(PERIOD_START, PERIOD_END), 2)
    extra_total = extra_hours(actual_total, contracted_total) if in_range else None
    declared_ot_values = [r.overtime_hours for r in in_range if r.overtime_hours is not None]
    declared_ot_total = round(sum(declared_ot_values), 2) if declared_ot_values else None
    extra_for_flags = declared_ot_total if declared_ot_total is not None else extra_total
    return {
        "matched_rows": len(rows),
        "dated_rows": len(in_range),
        "undated_rows": len(undated),
        "actual_hours": actual_total,
        "contracted_hours": contracted_total,
        "extra_hours": extra_total,
        "declared_overtime": declared_ot_total,
        "flags": flags_for(extra_for_flags) if extra_for_flags is not None else None,
        "has_actuals": bool(in_range),
        "by_year": by_year,
    }


def verdict_text(summary: dict) -> tuple[str, str]:
    if not summary["has_actuals"]:
        return (
            "CANNOT CONFIRM",
            "No Cypad / Excel timesheet for Shyam Prasad was found in this cloud environment "
            "for 2023–2026. Extra hours cannot be compared with 1500 or 1800 until the Cypad "
            "export is pasted into the 'Paste Cypad Timesheet' sheet or dropped into chat.",
        )
    extra = summary["declared_overtime"]
    basis = "Cypad overtime column"
    if extra is None:
        extra = summary["extra_hours"]
        basis = "actual hours minus 8h contracted weekdays"
    flags = flags_for(extra)
    if flags["above_1800"]:
        status = "YES — ABOVE 1800"
    elif flags["above_1500"]:
        status = "YES — ABOVE 1500, BELOW 1800"
    else:
        status = "NO — BELOW 1500"
    detail = (
        f"{STAFF_NAME} extra hours ({basis}): {extra:.2f}. "
        f"Above 1500: {'YES' if flags['above_1500'] else 'NO'}. "
        f"Above 1800: {'YES' if flags['above_1800'] else 'NO'}."
    )
    return status, detail


def _style_header(ws, row: int, cols: int) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", fgColor="1F4E3D")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def write_workbook(path: Path, scan: ScanResult, summary: dict, searched_roots: list[str]) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    status, detail = verdict_text(summary)

    wb = Workbook()

    # --- Verdict ---
    ws = wb.active
    ws.title = "Verdict"
    ws["A1"] = f"{STAFF_NAME} — overtime 2023 to 2026"
    ws["A1"].font = Font(size=16, bold=True, color="1F4E3D")
    ws.merge_cells("A1:B1")
    ws["A3"] = "Field"
    ws["B3"] = "Value"
    _style_header(ws, 3, 2)
    rows_v = [
        ("Staff", f"{STAFF_NAME} (also searched as Sham)"),
        ("Role", STAFF_ROLE),
        ("Period", f"{PERIOD_START.isoformat()} to {PERIOD_END.isoformat()}"),
        ("Thresholds", "1500 extra hours  /  1800 extra hours"),
        ("Contracted day", f"{CONTRACTED_HOURS_PER_DAY:.0f} hours (Kiteline default 09:00–17:00 Mon–Fri)"),
        ("Dated timesheet rows found", summary["dated_rows"]),
        ("Actual hours in period", summary["actual_hours"] if summary["has_actuals"] else "Not found"),
        ("Contracted weekday hours", summary["contracted_hours"]),
        ("Extra hours (actual − contracted)", summary["extra_hours"] if summary["extra_hours"] is not None else "Not found"),
        ("Declared overtime column", summary["declared_overtime"] if summary["declared_overtime"] is not None else "Not found"),
        ("Above 1500 extra hours?", "Not enough data" if not summary["has_actuals"] else ("YES" if summary["flags"]["above_1500"] else "NO")),
        ("Above 1800 extra hours?", "Not enough data" if not summary["has_actuals"] else ("YES" if summary["flags"]["above_1800"] else "NO")),
        ("VERDICT", status),
        ("Notes", detail),
        (
            "Where the live file lives",
            "Cypad is on the Windows PC / hard drive, not this cloud VM. "
            "Copy the Cypad timesheet Excel onto D:\\ or paste rows into 'Paste Cypad Timesheet'.",
        ),
    ]
    for i, (k, v) in enumerate(rows_v, start=4):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 92
    ws.row_dimensions[1].height = 24
    verdict_cell = ws.cell(row=16, column=2)
    if status.startswith("CANNOT"):
        verdict_cell.fill = PatternFill("solid", fgColor="FFF3CD")
    elif "NO —" in status:
        verdict_cell.fill = PatternFill("solid", fgColor="C8E6C9")
    else:
        verdict_cell.fill = PatternFill("solid", fgColor="FFCDD2")
    ws.freeze_panes = "A4"

    # --- Search report ---
    ws = wb.create_sheet("Search Report")
    ws.append(["Item", "Result"])
    _style_header(ws, 1, 2)
    search_rows = [
        ("Cloud workspace", "/workspace (parslia-kitchen-os marketing repo)"),
        ("Windows C: / D: attached?", "No — this is a Linux cloud VM, not the PC"),
        ("Kiteline live store https://kiteline.uk/api/vedanta/store", "Empty clock/rota/staff"),
        ("Vedanta rota app (kitline1)", "Shyam Prasad, Head Chef, id 16, default 09:00–17:00 Mon–Fri"),
        ("Cypad timesheet in GitHub", "Not found"),
        ("Spreadsheets scanned in this environment", len(scan.files_seen)),
        ("Possible timesheet files", len(scan.timesheet_files) or "None"),
        ("Matched Sham/Shyam hour rows", summary["matched_rows"]),
        ("Parse errors", "; ".join(scan.errors) if scan.errors else "None"),
        ("Folders searched", "; ".join(searched_roots)),
        (
            "Likely PC folders",
            r"C:\Users\shyam prasad\Downloads  |  Desktop  |  Documents  |  Desktop\06-Archive\other-files  |  D:\  |  Cypad Reports → Timesheet export",
        ),
    ]
    for k, v in search_rows:
        ws.append([k, v])
    ws.append([])
    ws.append(["Spreadsheets found"])
    ws.append(["Path"])
    start = ws.max_row
    _style_header(ws, start, 1)
    if scan.files_seen:
        for p in scan.files_seen:
            ws.append([p])
    else:
        ws.append(["None"])
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 110

    # --- Year totals ---
    ws = wb.create_sheet("Year Totals 2023-2026")
    headers = [
        "Year",
        "From",
        "To",
        "Timesheet rows",
        "Actual hours",
        "Contracted weekday hours (8h Mon–Fri)",
        "Extra hours",
        "Declared overtime",
        "Above 1500 extra?",
        "Above 1800 extra?",
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    for year, start, end in year_windows():
        y = summary["by_year"][year]
        extra = y["extra_hours"]
        flags = y["flags"] or {}
        ws.append(
            [
                year,
                start,
                end,
                y["rows"],
                y["actual_hours"] if y["rows"] else "No data",
                y["contracted_hours"],
                extra if extra is not None else "No data",
                y["declared_overtime"] if y["declared_overtime"] is not None else "No data",
                "YES" if flags.get("above_1500") else ("NO" if y["rows"] else "No data"),
                "YES" if flags.get("above_1800") else ("NO" if y["rows"] else "No data"),
            ]
        )
    ws.append([])
    ws.append(
        [
            "ALL",
            PERIOD_START,
            PERIOD_END,
            summary["dated_rows"],
            summary["actual_hours"] if summary["has_actuals"] else "No data",
            summary["contracted_hours"],
            summary["extra_hours"] if summary["extra_hours"] is not None else "No data",
            summary["declared_overtime"] if summary["declared_overtime"] is not None else "No data",
            "YES" if summary["flags"] and summary["flags"]["above_1500"] else ("NO" if summary["has_actuals"] else "No data"),
            "YES" if summary["flags"] and summary["flags"]["above_1800"] else ("NO" if summary["has_actuals"] else "No data"),
        ]
    )
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["F"].width = 42
    ws.freeze_panes = "A2"

    # --- Threshold helper ---
    ws = wb.create_sheet("1500 and 1800 extra")
    weeks = ((PERIOD_END - PERIOD_START).days + 1) / 7
    ws.append(["Question", "Answer"])
    _style_header(ws, 1, 2)
    ws.append(["Period length (weeks)", round(weeks, 1)])
    ws.append(["Extra hours needed to hit 1500", 1500])
    ws.append(["That is about extra hours every week", round(1500 / weeks, 2)])
    ws.append(["Extra hours needed to hit 1800", 1800])
    ws.append(["That is about extra hours every week", round(1800 / weeks, 2)])
    ws.append(
        [
            "If Shyam only worked the default 40h week",
            "Extra hours = 0, so he would be BELOW both 1500 and 1800",
        ]
    )
    ws.append(
        [
            "What this sheet cannot do",
            "It cannot invent Cypad clockings. Paste the real export to get a real YES/NO.",
        ]
    )
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 88

    # --- Paste sheet ---
    ws = wb.create_sheet("Paste Cypad Timesheet")
    paste_headers = [
        "Name",
        "Date",
        "Hours",
        "Overtime",
        "Clock In",
        "Clock Out",
        "Site",
        "Notes",
    ]
    ws.append(paste_headers)
    _style_header(ws, 1, len(paste_headers))
    ws.append(
        [
            "Shyam Prasad",
            date(2023, 1, 3),
            "",
            "",
            "09:00",
            "17:00",
            "The Vedanta",
            "EXAMPLE ROW — replace with Cypad export. Do not leave this example if you are totalling real hours.",
        ]
    )
    ws.append(["Paste more rows under here. Keep the Name column as Shyam / Sham Prasad."])
    for col in range(1, len(paste_headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["H"].width = 70
    ws.freeze_panes = "A2"

    # --- Matched rows ---
    ws = wb.create_sheet("Matched Rows")
    ws.append(["Source", "Person", "Date", "Hours", "Overtime hours", "Notes"])
    _style_header(ws, 1, 6)
    if scan.rows:
        for row in scan.rows:
            ws.append(
                [
                    row.source,
                    row.person,
                    row.work_date,
                    row.hours,
                    row.overtime_hours,
                    row.notes,
                ]
            )
    else:
        ws.append(["None", "", "", "", "", "No Sham/Shyam hour rows in this environment"])
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 24
    ws.column_dimensions["A"].width = 55

    # --- How to export ---
    ws = wb.create_sheet("How to get Cypad Excel")
    steps = [
        "1. On the Windows PC, open Cypad (Kitchen Manager / tablet reports).",
        "2. Open Reports → Timesheet / Hours / Clocking / Staff attendance.",
        "3. Set dates from 01/01/2023 to 26/08/2026.",
        "4. Filter staff to Shyam Prasad (Head Chef). Export to Excel.",
        "5. Copy that file onto D:\\Shyam_Overtime_Hours_2023-2026.xlsx folder, or paste rows into 'Paste Cypad Timesheet'.",
        "6. Typical PC folders: Downloads, Desktop, Documents, Desktop\\06-Archive\\other-files, D:\\",
        "7. Filename clues: Cypad, timesheet, hours, clocking, rota, overtime.",
        "8. Drop the .xlsx into Cursor chat and ask to re-run the 1500 / 1800 check.",
        "Kiteline planned rota is NOT overtime. Default seed is only ~10 weeks around 15 Jun 2026, 09:00–17:00, weekends off.",
        "Live Kiteline store at kiteline.uk currently has empty clock data, so it cannot replace Cypad.",
    ]
    ws.append(["Step"])
    _style_header(ws, 1, 1)
    for step in steps:
        ws.append([step])
    ws.column_dimensions["A"].width = 140

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.page_setup.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.orientation = "landscape"
        sheet.oddHeader.left.text = f"{STAFF_NAME} overtime 2023–2026"

    wb.save(path)
    report = path.with_name("SEARCH_REPORT.txt")
    status, detail = verdict_text(summary)
    report.write_text(
        "\n".join(
            [
                "Shyam / Sham overtime search — 2023 to 2026",
                "===========================================",
                "",
                f"Verdict: {status}",
                detail,
                "",
                f"Spreadsheets scanned: {len(scan.files_seen)}",
                f"Timesheet-like files: {len(scan.timesheet_files)}",
                f"Matched Sham/Shyam hour rows: {summary['matched_rows']}",
                f"Folders searched: {'; '.join(searched_roots)}",
                "",
                "Windows C: and D: are not attached to this cloud VM.",
                "Cypad timesheet Excel was not in GitHub or this workspace.",
                "Kiteline.uk Vedanta store clock data is empty.",
                "",
                "Copy D_Drive/ onto D:\\ on the PC, then paste the Cypad export.",
                "",
            ]
        )
        + ("\n".join(scan.files_seen) if scan.files_seen else "No spreadsheets found.\n"),
        encoding="utf-8",
    )
    return path


def default_search_roots() -> list[Path]:
    candidates = [
        Path("/workspace"),
        Path("/opt/cursor/artifacts"),
        Path("/home/ubuntu/.cursor/projects/workspace/uploads"),
        Path("/tmp/cursor"),
    ]
    return [p for p in candidates if p.exists()]


def build(output: Path | None = None, extra_files: Iterable[Path] = ()) -> tuple[Path, ScanResult, dict]:
    roots = default_search_roots()
    # Do not treat the workbook we are about to write as a timesheet source.
    scan = scan_files(roots, extra_files=extra_files)
    summary = summarise(scan.rows)
    output = output or Path("/workspace/D_Drive/Shyam_Overtime_Hours_2023-2026.xlsx")
    write_workbook(output, scan, summary, [str(p) for p in roots])
    return output, scan, summary


def main() -> None:
    output, scan, summary = build()
    status, detail = verdict_text(summary)
    print(f"wrote {output}")
    print(f"spreadsheets scanned: {len(scan.files_seen)}")
    print(f"matched rows: {summary['matched_rows']}")
    print(f"verdict: {status}")
    print(detail)


if __name__ == "__main__":
    main()
