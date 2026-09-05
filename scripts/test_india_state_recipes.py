#!/usr/bin/env python3
"""Check Indian state recipe folders and Excel workbooks."""

from __future__ import annotations

import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "recipes" / "onion-garlic-free-indian" / "india-states"
COURSES = [
    "01-starters",
    "02-mains",
    "03-sides",
    "04-breads",
    "05-sweets",
    "06-desserts",
    "07-salads",
]


def fail(msg: str) -> None:
    print(f"FAIL: {msg}")
    sys.exit(1)


def main() -> None:
    if not OUT.is_dir():
        fail(f"missing {OUT}")
    states = sorted(
        p for p in OUT.iterdir() if p.is_dir() and p.name[:2].isdigit()
    )
    if len(states) != 36:
        fail(f"expected 36 state folders, got {len(states)}")
    for folder in states:
        if not (folder / "README.md").is_file():
            fail(f"missing README in {folder.name}")
        xlsx = list((folder / "excel").glob("*-recipes.xlsx"))
        if len(xlsx) != 1:
            fail(f"{folder.name} workbooks: {xlsx}")
        for course in COURSES:
            md = list((folder / course).glob("*.md"))
            if len(md) != 1:
                fail(f"{folder.name}/{course}: {md}")

    master = OUT / "excel" / "ALL-STATES.xlsx"
    shopping = OUT / "excel" / "SHOPPING-LIST.xlsx"
    if not master.is_file():
        fail("missing ALL-STATES.xlsx")
    if not shopping.is_file():
        fail("missing SHOPPING-LIST.xlsx")

    wb = load_workbook(master, read_only=True, data_only=True)
    required = {
        "Cover",
        "Rules",
        "All Recipes",
        "North",
        "South",
        "East",
        "West",
        "Central",
        "Northeast",
        "Union Territory",
        "Starters",
        "Mains",
        "Sides",
        "Breads",
        "Sweets",
        "Desserts",
        "Salads",
        "Shopping list",
    }
    missing = required - set(wb.sheetnames)
    if missing:
        fail(f"master missing sheets: {missing}")

    rows = []
    for row in wb["All Recipes"].iter_rows(min_row=6, values_only=True):
        if row[3]:
            rows.append((row[0], row[2], row[3], row[11]))
    if len(rows) != 252:
        fail(f"All Recipes has {len(rows)} dishes, expected 252")
    if set(Counter(r[0] for r in rows).values()) != {7}:
        fail(f"per-state counts {Counter(r[0] for r in rows)}")
    if set(Counter(r[1] for r in rows).values()) != {36}:
        fail(f"per-course counts {Counter(r[1] for r in rows)}")
    for _state, _cat, name, diet in rows:
        diet_l = (diet or "").lower()
        if "no onion" not in diet_l or "no garlic" not in diet_l:
            fail(f"{name} diet: {diet}")
    banner = wb["Cover"]["A3"].value
    if banner is None or "NO ONION" not in str(banner).upper():
        fail(f"cover banner: {banner!r}")
    shop_rows = sum(1 for row in wb["Shopping list"].iter_rows(min_row=6, values_only=True) if row[0])
    wb.close()
    if shop_rows < 1000:
        fail(f"shopping list too short: {shop_rows}")

    md_count = len(list(OUT.rglob("*.md")))
    xlsx_count = len(list(OUT.rglob("*.xlsx")))
    if md_count != 289:
        fail(f"expected 289 markdown files, got {md_count}")
    if xlsx_count != 38:
        fail(f"expected 38 Excel workbooks, got {xlsx_count}")

    print("PASS")
    print("  recipes: 252 (36 states/UTs x 7 courses)")
    print(f"  markdown: {md_count}")
    print(f"  excel workbooks: {xlsx_count}")
    print(f"  shopping-list rows: {shop_rows}")


if __name__ == "__main__":
    main()
