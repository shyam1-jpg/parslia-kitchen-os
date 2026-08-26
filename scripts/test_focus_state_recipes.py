#!/usr/bin/env python3
"""Check deep-kitchen (focus-state) recipe folders and Excel workbooks."""

from __future__ import annotations

import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "recipes" / "onion-garlic-free-indian" / "focus-states"
COURSES = [
    "01-starters",
    "02-mains",
    "03-sides",
    "04-breads",
    "05-sweets",
    "06-desserts",
    "07-salads",
]
REQUIRED_KITCHENS = {
    "14-rajasthan",
    "15-gujarat",
    "16-punjab",
    "17-pan-india",
}
KITCHEN_COUNT = 17
RECIPES_EACH = 21
EXPECTED_RECIPES = KITCHEN_COUNT * RECIPES_EACH  # 357
EXPECTED_MD = 1 + KITCHEN_COUNT + EXPECTED_RECIPES  # cover + kitchen readmes + recipes
EXPECTED_XLSX = 2 + KITCHEN_COUNT  # master + shopping + per kitchen


def fail(msg: str) -> None:
    print(f"FAIL: {msg}")
    sys.exit(1)


def main() -> None:
    if not OUT.is_dir():
        fail(f"missing {OUT}")
    kitchens = sorted(p for p in OUT.iterdir() if p.is_dir() and p.name[:2].isdigit())
    if len(kitchens) != KITCHEN_COUNT:
        fail(f"expected {KITCHEN_COUNT} kitchen folders, got {len(kitchens)}")
    names = {p.name for p in kitchens}
    missing_k = REQUIRED_KITCHENS - names
    if missing_k:
        fail(f"missing requested kitchens: {missing_k}")
    for folder in kitchens:
        if not (folder / "README.md").is_file():
            fail(f"missing README in {folder.name}")
        xlsx = list((folder / "excel").glob("*-recipes.xlsx"))
        if len(xlsx) != 1:
            fail(f"{folder.name} workbooks: {xlsx}")
        for course in COURSES:
            md = list((folder / course).glob("*.md"))
            if len(md) != 3:
                fail(f"{folder.name}/{course}: expected 3 recipes, got {md}")
            html = list((folder / course).glob("*.html"))
            if len(html) != 3:
                fail(f"{folder.name}/{course}: expected 3 HTML cards, got {html}")
        if not (folder / "index.html").is_file():
            fail(f"missing recipe-card index {folder.name}/index.html")

    master = OUT / "excel" / "FOCUS-STATES.xlsx"
    shopping = OUT / "excel" / "SHOPPING-LIST.xlsx"
    if not master.is_file():
        fail("missing FOCUS-STATES.xlsx")
    if not shopping.is_file():
        fail("missing SHOPPING-LIST.xlsx")

    wb = load_workbook(master, read_only=True, data_only=True)
    required_sheets = {
        "Cover",
        "Rules",
        "All Recipes",
        "Rajasthan",
        "Gujarat",
        "Punjab",
        "Pan-India",
        "Starters",
        "Mains",
        "Sides",
        "Breads",
        "Sweets",
        "Desserts",
        "Salads",
        "Shopping list",
    }
    missing = required_sheets - set(wb.sheetnames)
    if missing:
        fail(f"master missing sheets: {missing}")

    rows = []
    for row in wb["All Recipes"].iter_rows(min_row=6, values_only=True):
        if row[3]:
            rows.append((row[0], row[2], row[3], row[7], row[8], row[9], row[11]))
    if len(rows) != EXPECTED_RECIPES:
        fail(f"All Recipes has {len(rows)} dishes, expected {EXPECTED_RECIPES}")
    if set(Counter(r[0] for r in rows).values()) != {RECIPES_EACH}:
        fail(f"per-kitchen counts {Counter(r[0] for r in rows)}")
    if set(Counter(r[1] for r in rows).values()) != {KITCHEN_COUNT * 3}:
        fail(f"per-course counts {Counter(r[1] for r in rows)}")

    heartland = {"Rajasthan", "Gujarat", "Punjab", "Pan-India"}
    if heartland - {r[0] for r in rows}:
        fail(f"All Recipes missing kitchens: {heartland - {r[0] for r in rows}}")

    for kitchen, _cat, name, cookware, _ings, _method, diet in rows:
        diet_l = (diet or "").lower()
        cook_l = (cookware or "").lower()
        if "no onion" not in diet_l or "no garlic" not in diet_l:
            fail(f"{kitchen}/{name} diet: {diet}")
        if "aluminium" not in cook_l and "aluminum" not in cook_l and "steel" not in cook_l and "iron" not in cook_l and "glass" not in cook_l and "clay" not in cook_l:
            fail(f"{kitchen}/{name} cookware looks unsafe: {cookware}")
        if "aluminium" in cook_l and "never" not in cook_l and "not" not in cook_l:
            fail(f"{kitchen}/{name} aluminium without never/not: {cookware}")

    banner = wb["Cover"]["A3"].value
    if banner is None or "NO ONION" not in str(banner).upper():
        fail(f"cover banner: {banner!r}")
    shop_rows = sum(1 for row in wb["Shopping list"].iter_rows(min_row=6, values_only=True) if row[0])
    wb.close()
    if shop_rows < 400:
        fail(f"shopping list too short: {shop_rows}")

    md_count = len(list(OUT.rglob("*.md")))
    xlsx_count = len(list(OUT.rglob("*.xlsx")))
    html_count = len(list(OUT.rglob("*.html")))
    if md_count != EXPECTED_MD:
        fail(f"expected {EXPECTED_MD} markdown files, got {md_count}")
    if xlsx_count != EXPECTED_XLSX:
        fail(f"expected {EXPECTED_XLSX} Excel workbooks, got {xlsx_count}")
    if html_count != EXPECTED_RECIPES + KITCHEN_COUNT + 2:
        fail(f"expected {EXPECTED_RECIPES + KITCHEN_COUNT + 2} HTML files, got {html_count}")
    if not (OUT / "DOWNLOAD.html").is_file():
        fail("missing DOWNLOAD.html")
    zips = list((OUT / "download").glob("*.zip"))
    if len(zips) != KITCHEN_COUNT + 1:
        fail(f"expected {KITCHEN_COUNT + 1} zip packs, got {zips}")
    if not (OUT / "download" / "07-manipur.zip").is_file():
        fail("missing Manipur ZIP pack")
    if not (OUT / "download" / "ALL-FOCUS-KITCHEN-RECIPES.zip").is_file():
        fail("missing all-kitchens ZIP")
    manipur_pdfs = list((OUT / "07-manipur").rglob("*.pdf"))
    if len(manipur_pdfs) != RECIPES_EACH:
        fail(f"Manipur PDF cards: {len(manipur_pdfs)}")
    chamthong = OUT / "07-manipur" / "02-mains" / "chamthong-stew.pdf"
    if not chamthong.is_file() or chamthong.stat().st_size < 1000:
        fail(f"missing downloadable Chamthong PDF: {chamthong}")
    booklet = OUT / "download" / "07-manipur-recipes.pdf"
    if not booklet.is_file() or booklet.stat().st_size < 5000:
        fail("missing Manipur 21-card PDF booklet")

    # Professional card: Manipur workbook has Qty/Unit/Ingredient and measured lines.
    manipur_xlsx = OUT / "07-manipur" / "excel" / "manipur-recipes.xlsx"
    mwb = load_workbook(manipur_xlsx, read_only=True, data_only=True)
    if "Menu" not in mwb.sheetnames:
        fail("Manipur workbook missing Menu sheet")
    card_sheets = [n for n in mwb.sheetnames if n not in {"Menu", "Rules", "Shopping list", "All dishes"}]
    if len(card_sheets) != RECIPES_EACH:
        fail(f"Manipur recipe cards: {len(card_sheets)} sheets, expected {RECIPES_EACH}")
    sample = mwb[card_sheets[3]]  # a main, Chamthong is 4th recipe
    header_vals = []
    for row in sample.iter_rows(min_row=1, max_row=16, max_col=4, values_only=True):
        header_vals.extend([str(v) for v in row if v])
    blob = " ".join(header_vals).lower()
    if "qty" not in blob or "unit" not in blob or "ingredient" not in blob:
        fail(f"Manipur card missing Qty/Unit/Ingredient headers: {header_vals[:20]}")
    if "recipe card" not in blob:
        fail("Manipur sheet is not titled as a recipe card")
    mwb.close()

    from focus_state_recipe_data import FOCUS_STATES, build_recipes
    import build_continent_recipes as core

    recipes = build_recipes(core.r)
    missing_qty = []
    for rec in recipes:
        for item in rec["ingredients"]:
            parsed = item if isinstance(item, dict) else {"qty": ""}
            if not str(parsed.get("qty") or "").strip():
                missing_qty.append(f"{rec['continent_id']}/{rec['name']}: {parsed}")
    if missing_qty:
        fail(f"{len(missing_qty)} ingredients have no quantity, e.g. {missing_qty[:8]}")

    print("PASS")
    print(f"  recipes: {EXPECTED_RECIPES} ({KITCHEN_COUNT} kitchens x {RECIPES_EACH})")
    print(f"  markdown: {md_count}")
    print(f"  html recipe cards: {html_count}")
    print(f"  excel workbooks: {xlsx_count}")
    print(f"  shopping-list rows: {shop_rows}")
    print("  includes: Rajasthan, Gujarat, Punjab, Pan-India")
    print("  Manipur: folder + workbook + 21 measured recipe cards")


if __name__ == "__main__":
    main()
