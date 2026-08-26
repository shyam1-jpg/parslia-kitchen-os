#!/usr/bin/env python3
"""Build onion-garlic-free vegetarian Indian recipes by continent.

Creates:
  recipes/onion-garlic-free-indian/
    README.md
    COOKWARE-AND-DIET-RULES.md
    excel/ALL-CONTINENTS.xlsx
    excel/SHOPPING-LIST.xlsx
    <continent>/README.md
    <continent>/excel/<continent>-recipes.xlsx
    <continent>/<category>/<recipe>.md
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.child import INVALID_TITLE_REGEX
import re

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "recipes" / "onion-garlic-free-indian"

# --- style ---
SAFFRON = PatternFill("solid", fgColor="9A3412")
CREAM = PatternFill("solid", fgColor="FFF7ED")
SAND = PatternFill("solid", fgColor="FFEDD5")
GREEN = PatternFill("solid", fgColor="14532D")
GREEN_SOFT = PatternFill("solid", fgColor="DCFCE7")
RED = PatternFill("solid", fgColor="7F1D1D")
RED_SOFT = PatternFill("solid", fgColor="FEE2E2")
GOLD = PatternFill("solid", fgColor="B45309")
WHITE = PatternFill("solid", fgColor="FFFFFF")
SLATE = PatternFill("solid", fgColor="1E293B")
CATEGORY_FILL = {
    "Starter": PatternFill("solid", fgColor="FDBA74"),
    "Main": PatternFill("solid", fgColor="FB923C"),
    "Side": PatternFill("solid", fgColor="86EFAC"),
    "Bread": PatternFill("solid", fgColor="FDE68A"),
    "Sweet": PatternFill("solid", fgColor="F9A8D4"),
    "Dessert": PatternFill("solid", fgColor="C4B5FD"),
    "Salad": PatternFill("solid", fgColor="6EE7B7"),
}
THIN = Border(
    left=Side(style="thin", color="D6D3D1"),
    right=Side(style="thin", color="D6D3D1"),
    top=Side(style="thin", color="D6D3D1"),
    bottom=Side(style="thin", color="D6D3D1"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=18)
SUB_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
BODY_FONT = Font(name="Calibri", size=11, color="1C1917")
BOLD = Font(name="Calibri", bold=True, size=11, color="1C1917")
SMALL = Font(name="Calibri", size=10, italic=True, color="44403C")

CONTINENTS = [
    {
        "id": "asia",
        "folder": "01-asia",
        "name": "Asia",
        "community": "India — Gujarati, North Indian, South Indian",
        "note": "The home kitchen. These are the classic no-onion no-garlic dishes the rest of the world adapted.",
    },
    {
        "id": "africa",
        "folder": "02-africa",
        "name": "Africa",
        "community": "Durban Indian and East African Gujarati (Kenya, Uganda, Tanzania)",
        "note": "Bold chilli, curry leaves, coconut and brinjal. Onion and garlic removed; hing, ginger and tomato carry the gravy.",
    },
    {
        "id": "europe",
        "folder": "03-europe",
        "name": "Europe",
        "community": "British Indian restaurant cooking",
        "note": "Tikka masala, naan and paneer tikka rewritten without alliums for a sattvic Sunday roast-style spread.",
    },
    {
        "id": "north-america",
        "folder": "04-north-america",
        "name": "North America",
        "community": "Punjabi-Canadian and US Indian home cooking",
        "note": "Chana, saag, samosa and kulcha — the vegetarian backbone of Indian North America, onion-garlic free.",
    },
    {
        "id": "south-america",
        "folder": "05-south-america",
        "name": "South America & Caribbean",
        "community": "Trinidad, Guyana and Suriname Indo-Caribbean",
        "note": "Channa and aloo, sada roti, baigan choka and parsad. Geera and curry powder replace onion and garlic.",
    },
    {
        "id": "oceania",
        "folder": "06-oceania",
        "name": "Oceania",
        "community": "Indo-Fijian",
        "note": "Jackfruit khatar, dhal and roti from Fiji’s Girmitiya kitchens, cooked sattvic and aluminium-free.",
    },
    {
        "id": "antarctica",
        "folder": "07-antarctica",
        "name": "Antarctica",
        "community": "Polar / expedition kitchen (shelf-stable sattvic Indian)",
        "note": "No local farms. These recipes use dry, canned and long-life stores a research-station cook can actually pack.",
    },
]

CAT_FOLDERS = {
    "Starter": "01-starters",
    "Main": "02-mains",
    "Side": "03-sides",
    "Bread": "04-breads",
    "Sweet": "05-sweets",
    "Dessert": "06-desserts",
    "Salad": "07-salads",
}

RULES_MD = """# Diet and cookware rules

Every recipe in this collection follows the same kitchen rules.

## Diet

- Vegetarian (no meat, fish, eggs)
- **No onion, garlic, shallot, leek, chive, spring onion**
- No onion powder or garlic powder (read curry-powder labels)
- Ginger, hing (asafoetida), tomato, coconut, yogurt and whole spices are allowed
- Hing in hot ghee or oil is the stand-in for onion and garlic — use a **pinch** only

## Cookware — no aluminium

Acidic Indian food (tomato, tamarind, lemon, yogurt) reacts with aluminium.
It can pick up a metallic taste and leach metal into the food.

**Use**
- Stainless steel pots, kadhai, steamers and thalis
- Cast-iron tawa for breads
- Clay / earthen pots (optional, for dal)
- Enamel, glass or ceramic for baking and serving

**Do not use**
- Aluminium pots, steamers or thalis
- Aluminium foil trays
- Aluminium pressure-cooker inners for tomato gravies if a steel inner is available

## Flavour map (instead of onion and garlic)

| Need | Use |
|------|-----|
| Savoury depth | Pinch of hing in hot fat |
| Fresh heat | Ginger + green chilli |
| Body in gravy | Tomato puree + cashew paste |
| Restaurant finish | Kasuri methi + garam masala + cream |
| South Indian snap | Mustard seeds, urad dal, curry leaves, coconut |
| Tang | Lemon, amchur, tamarind, yogurt |
"""

README_MD = """# Onion-garlic-free vegetarian Indian recipes — all continents

Vegetarian Indian cooking as it travelled with the diaspora, rewritten **without onion or garlic** and cooked **without aluminium**.

Each continent folder has:

- a `README.md` menu
- one markdown file per recipe, grouped by course
- an Excel workbook in `excel/`

Open the master spreadsheet first:

- [`excel/ALL-CONTINENTS.xlsx`](excel/ALL-CONTINENTS.xlsx) — every recipe, one sheet per continent, plus a filterable All Recipes table
- [`excel/SHOPPING-LIST.xlsx`](excel/SHOPPING-LIST.xlsx) — combined shopping list

## Diet flags

Vegetarian · no onion · no garlic · no other alliums · no aluminium cookware

See [COOKWARE-AND-DIET-RULES.md](COOKWARE-AND-DIET-RULES.md).

## Folders

| Folder | Continent | Kitchen |
|--------|-----------|---------|
| `01-asia` | Asia | India |
| `02-africa` | Africa | Durban + East African Gujarati |
| `03-europe` | Europe | British Indian |
| `04-north-america` | North America | Punjabi-Canadian / US |
| `05-south-america` | South America & Caribbean | Trinidad, Guyana, Suriname |
| `06-oceania` | Oceania | Indo-Fijian |
| `07-antarctica` | Antarctica | Polar / expedition stores |

## Courses in every folder

Starter · Main · Side · Bread · Sweet · Dessert · Salad

Serves 4 unless a recipe says otherwise.

## Indian states (28 + 8 union territories)

A full seven-course vegetarian menu for **every Indian state and UT** is in [`india-states/`](india-states/README.md).

- [`india-states/excel/ALL-STATES.xlsx`](india-states/excel/ALL-STATES.xlsx) — 252 recipes
- [`india-states/excel/SHOPPING-LIST.xlsx`](india-states/excel/SHOPPING-LIST.xlsx)

## Focus kitchens (Rajasthan, Gujarat, Punjab, pan-India, and more)

A **21-recipe** vegetarian library for each requested kitchen is in [`focus-states/`](focus-states/README.md).

- [`focus-states/excel/FOCUS-STATES.xlsx`](focus-states/excel/FOCUS-STATES.xlsx) — 357 recipes
- [`focus-states/excel/SHOPPING-LIST.xlsx`](focus-states/excel/SHOPPING-LIST.xlsx)

Includes **Rajasthan**, **Gujarat**, **Punjab**, and **pan-India** dishes eaten all over the country, plus Goa, Maharashtra, Odisha, Bengal, Andhra, Kerala, Manipur, Meghalaya, Uttarakhand, UP, Bihar, Mithila, and Karnataka.

## Rebuild

```bash
python3 scripts/build_continent_recipes.py
python3 scripts/build_india_state_recipes.py
python3 scripts/build_focus_state_recipes.py
```
"""


def slug(text: str) -> str:
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def sheet_title(text: str) -> str:
    text = INVALID_TITLE_REGEX.sub(" ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:31]


def r(continent_id, category, name, community, prep, cook, cookware, why, ingredients, method, notes="", servings=4):
    return {
        "continent_id": continent_id,
        "category": category,
        "name": name,
        "community": community,
        "servings": servings,
        "prep_min": prep,
        "cook_min": cook,
        "cookware": cookware,
        "why": why,
        "ingredients": ingredients,
        "method": method,
        "notes": notes,
        "diet": "Vegetarian · no onion · no garlic · no allium · no aluminium",
    }


def all_recipes() -> list[dict]:
    S = "Stainless steel kadhai or saucepan — never aluminium"
    STEAM = "Stainless steel steamer and steel thali — never aluminium"
    TAWA = "Cast-iron tawa (not aluminium)"
    OVEN = "Steel or ceramic oven tray — never an aluminium foil tray"
    GRILL = "Cast-iron grill pan or steel oven tray"
    CLAY = "Stainless steel or clay pot — never aluminium"

    return [
        # ---------- ASIA ----------
        r("asia", "Starter", "Khaman Dhokla", "Gujarat, India", 20, 15, STEAM,
          "The best Indian vegetarian starter that is already onion-garlic free. Light, spongy, and party-proof.",
          ["1½ cups gram flour (besan)", "2 tbsp sooji (semolina)", "1 tsp ginger-green chilli paste",
           "1 tsp sugar", "½ tsp turmeric", "¾ tsp salt", "1 tbsp lemon juice", "¾ cup water",
           "1 tsp Eno fruit salt, added last", "2 tbsp oil", "1 tsp mustard seeds", "1 tsp sesame seeds",
           "8–10 curry leaves", "2 slit green chillies", "½ cup water + 2 tsp sugar + 1 tsp lemon (pouring syrup)",
           "2 tbsp grated coconut", "2 tbsp chopped coriander"],
          ["Mix besan, sooji, ginger-chilli, sugar, turmeric, salt, lemon and water to a thick pouring batter. Rest 10 minutes.",
           "Bring water to a boil in a steel steamer. Grease a steel thali.",
           "Stir Eno into the batter until frothy. Pour at once into the thali.",
           "Steam 12–15 minutes until a knife comes out clean. Cool 5 minutes and cut into squares.",
           "Temper mustard, sesame, curry leaves and chilli in oil. Add the sweet-sour water and pour over the dhokla.",
           "Garnish with coconut and coriander."],
          "Do not steam in an aluminium thali — the lemon and Eno react with the metal."),
        r("asia", "Main", "Shahi Matar Paneer", "North India", 15, 25, S,
          "The richest no-onion no-garlic gravy. Cashew and tomato replace the allium base.",
          ["2 tbsp ghee", "1 bay leaf", "4 green cardamom", "1-inch cinnamon", "4 cloves",
           "Pinch of hing", "3 ripe tomatoes, chopped", "12 cashews, soaked 20 minutes",
           "1 tsp ginger paste", "1 tsp Kashmiri chilli powder", "1 tsp coriander powder", "¼ tsp turmeric",
           "200 g paneer, cubed", "¾ cup green peas", "½ cup water", "2 tbsp cream",
           "½ tsp garam masala", "1 tsp kasuri methi, crushed", "Salt", "Pinch of sugar", "Coriander to finish"],
          ["Blend tomatoes and soaked cashews to a smooth paste.",
           "Heat ghee in a steel kadhai. Bloom whole spices, then a pinch of hing.",
           "Add the paste, ginger and ground spices. Cook 8–10 minutes until ghee separates.",
           "Add water, paneer and peas. Simmer 5 minutes.",
           "Stir in cream, garam masala, kasuri methi, salt and sugar. Garnish."],
          "If the gravy tastes sharp, cook 2 minutes more and add a teaspoon of cream — do not add onion."),
        r("asia", "Side", "Cabbage Poriyal", "Tamil Nadu, India", 10, 12, S,
          "Fast South Indian dry vegetable. Coconut and mustard seeds do the flavour work.",
          ["2 tbsp coconut oil", "1 tsp mustard seeds", "1 tsp urad dal", "8 curry leaves",
           "Pinch of hing", "2 dry red chillies", "4 cups shredded cabbage", "¼ tsp turmeric",
           "Salt", "¼ cup grated coconut", "Coriander"],
          ["Heat coconut oil. Splutter mustard, urad dal, curry leaves, chilli and hing.",
           "Add cabbage, turmeric and salt. Cover 6–8 minutes until just tender.",
           "Uncover, stir in coconut and coriander. Do not overcook — keep a little bite."]),
        r("asia", "Bread", "Phulka", "North India", 20, 15, TAWA,
          "Everyday wholewheat bread. Puffs only on a properly hot iron tawa, not aluminium.",
          ["2 cups atta (wholewheat flour)", "Warm water as needed", "Pinch of salt", "Ghee for brushing"],
          ["Knead a soft dough. Rest 15 minutes under a cloth.",
           "Divide into 10 balls. Roll thin rounds.",
           "Cook the first side on a hot iron tawa 20 seconds. Flip.",
           "When small bubbles appear, flip onto the open flame (or press with a cloth) until it puffs.",
           "Brush with ghee. Keep wrapped in a cotton cloth."],
          "Aluminium tawas do not hold heat — the phulka will not puff."),
        r("asia", "Sweet", "Coconut Ladoo", "India (coastal / festival)", 5, 12, S,
          "The easiest mithai that is naturally onion-garlic free.",
          ["2 cups fresh or desiccated coconut", "¾ cup condensed milk (or ½ cup jaggery + 2 tbsp milk)",
           "1 tsp ghee", "4 green cardamom, crushed", "Extra coconut for rolling"],
          ["Warm ghee in a steel kadhai. Add coconut and condensed milk (or jaggery and milk).",
           "Cook on low, stirring, until the mix leaves the sides.",
           "Add cardamom. Cool until handleable. Shape into ladoos and roll in coconut."]),
        r("asia", "Dessert", "Saffron Rice Kheer", "North India", 10, 45, "Heavy stainless steel milk pot — never aluminium",
          "Festival dessert. Milk cooked in aluminium tastes metallic — steel only.",
          ["¼ cup basmati rice, rinsed and soaked 20 minutes", "1 litre full-fat milk",
           "⅓ cup sugar", "8 saffron strands soaked in 1 tbsp warm milk",
           "4 green cardamom, crushed", "10 raisins", "10 slivered almonds"],
          ["Bring milk to a simmer in a heavy steel pot.",
           "Add drained rice. Cook on low 35–40 minutes, stirring often, until creamy.",
           "Add sugar, saffron milk, cardamom, raisins and almonds. Cook 5 minutes more.",
           "Serve warm or chilled."],
          "Never reduce milk in aluminium — it greys the kheer and dulls saffron."),
        r("asia", "Salad", "Moong Kosambari", "Karnataka, India", 30, 2, "Steel mixing bowl + small steel tadka pan",
          "India’s best salad — already sattvic. Protein from soaked moong.",
          ["¼ cup split yellow moong, soaked 30 minutes and drained", "½ cucumber, diced",
           "2 tbsp grated coconut", "2 tbsp coriander, chopped", "1 tbsp lemon juice", "Salt",
           "1 tsp oil", "½ tsp mustard seeds", "6 curry leaves", "Pinch of hing", "1 dry red chilli"],
          ["Mix moong, cucumber, coconut, coriander, lemon and salt.",
           "Temper mustard, curry leaves, hing and chilli in oil. Pour over the salad. Toss."]),

        # ---------- AFRICA ----------
        r("africa", "Starter", "Cabbage Bhajia", "Durban Indian", 15, 15, "Steel kadhai for frying — never aluminium",
          "Durban’s rainy-day snack. Cabbage stands in for onion bhajia.",
          ["2 cups shredded cabbage", "1 cup gram flour", "2 tbsp rice flour",
           "1 tsp ginger-green chilli paste", "½ tsp turmeric", "1 tsp coriander powder",
           "½ tsp cumin seeds", "Pinch of hing", "Salt", "Water as needed", "Oil for frying (steel kadhai)"],
          ["Mix cabbage with salt and spices. Rest 5 minutes so it weeps a little moisture.",
           "Add flours and a splash of water to a thick coating batter.",
           "Fry spoonfuls in medium-hot oil until golden. Drain on a steel rack.",
           "Serve with tamarind-date chutney (no garlic)."]),
        r("africa", "Main", "Durban Brinjal Coconut Curry", "Durban / East African Indian", 15, 25, S,
          "The vegetarian curry of Durban Indian homes. Brinjal melts and thickens the gravy so you never miss onion.",
          ["4 tbsp oil", "1 tsp mustard seeds", "1 tsp cumin seeds", "10 curry leaves", "Pinch of hing",
           "4 small brinjals, quartered", "1 large tomato, blended", "1 tsp ginger paste",
           "1 tsp Kashmiri chilli powder", "½ tsp turmeric", "2 tsp coriander powder",
           "1 cup coconut milk", "Salt", "½ tsp sugar", "Lemon juice", "Coriander"],
          ["Heat oil. Splutter mustard, cumin, curry leaves and hing.",
           "Add brinjal and fry 4 minutes until the edges brown.",
           "Add tomato, ginger and ground spices. Cook until the oil shines.",
           "Pour coconut milk and salt. Simmer 15 minutes until glossy and thick.",
           "Finish with sugar, lemon and coriander."],
          "Use garlic-free curry powder if you add any. Many Durban mixes hide garlic."),
        r("africa", "Side", "Green Bean and Potato Curry", "East African Gujarati", 10, 20, S,
          "Weeknight sabzi from Kenyan and Ugandan Gujarati kitchens.",
          ["2 tbsp oil", "1 tsp cumin seeds", "Pinch of hing", "2 cups green beans, cut",
           "2 potatoes, cubed", "1 tomato, chopped", "½ tsp turmeric", "1 tsp coriander powder",
           "½ tsp chilli powder", "Salt", "½ cup water", "Lemon", "Coriander"],
          ["Temper cumin and hing in oil.",
           "Add potato, beans, tomato and spices. Stir 2 minutes.",
           "Add water, cover, cook 12–15 minutes until potatoes are soft.",
           "Crush a few potato cubes into the gravy. Finish with lemon."]),
        r("africa", "Bread", "Soft Durban Roti", "Durban Indian", 25, 20, TAWA,
          "The bread Durban curry is eaten with. Cooked on iron, not aluminium.",
          ["2 cups atta", "2 tbsp oil", "Pinch of salt", "Warm water", "Ghee for brushing"],
          ["Knead a soft, slightly oily dough. Rest 20 minutes.",
           "Roll medium-thick rounds. Cook both sides on a hot iron tawa.",
           "Smear ghee and stack in a cloth. Serve with brinjal curry."]),
        r("africa", "Sweet", "Coconut Cardamom Burfi", "East African Indian", 10, 20, S,
          "Coastal East African Indian mithai. Coconut, milk and cardamom only.",
          ["2 cups grated coconut", "1 cup full-fat milk", "¾ cup sugar", "1 tsp ghee",
           "4 cardamom pods, crushed", "Pistachios to garnish"],
          ["Cook coconut, milk and sugar in a steel kadhai on low until thick.",
           "Add ghee and cardamom. Cook until it leaves the sides.",
           "Press into a greased steel or glass tray. Top with pistachios. Cool and cut."]),
        r("africa", "Dessert", "Semiya Payasam", "East African / South Indian diaspora", 5, 20, S,
          "Vermicelli milk pudding served after East African Indian thalis.",
          ["1 tbsp ghee", "½ cup roasted vermicelli", "3 cups milk", "⅓ cup sugar",
           "8 saffron strands", "4 cardamom", "10 cashews", "10 raisins"],
          ["Fry cashews and raisins in ghee; set aside. In the same ghee toast vermicelli until golden.",
           "Add milk and simmer 10–12 minutes until the semiya is soft.",
           "Add sugar, saffron and cardamom. Cook 3 minutes. Top with nuts."],
          "Steel saucepan only — milk plus aluminium tastes of the pan."),
        r("africa", "Salad", "Tomato Cucumber Limbu Salad", "Durban", 10, 0, "Steel or glass bowl",
          "No-onion kachumber with extra lemon, the way Durban tables serve it next to hot curry.",
          ["2 tomatoes, diced", "1 cucumber, diced", "1 green chilli, sliced", "Coriander",
           "Juice of 1 lemon", "½ tsp roasted cumin", "Black salt", "Pinch of sugar"],
          ["Toss everything just before serving so the cucumber stays crisp.",
           "Do not add onion — chilli and lemon give the bite."]),

        # ---------- EUROPE ----------
        r("europe", "Starter", "Tandoori Paneer Tikka", "British Indian", 40, 15, GRILL,
          "The UK’s favourite Indian starter, marinade rewritten without garlic ginger paste full of garlic.",
          ["300 g paneer, cubed", "1 capsicum, chunks", "1 tomato, petals (seeds removed)",
           "200 g hung yogurt", "1 tsp ginger paste", "½ tsp turmeric", "1 tsp Kashmiri chilli powder",
           "1 tsp coriander powder", "½ tsp garam masala", "1 tsp kasuri methi",
           "1 tbsp lemon juice", "1 tbsp mustard oil", "Pinch of hing", "Salt", "Chaat masala to finish"],
          ["Mix the marinade. Coat paneer and vegetables. Rest 30 minutes.",
           "Skewer. Roast on a cast-iron grill or steel tray at 220°C for 12–15 minutes, turning once.",
           "Finish with lemon and chaat masala."],
          "Skip bottled tikka paste — almost all of them list garlic."),
        r("europe", "Main", "Vegetable Tikka Masala", "United Kingdom", 20, 30, S,
          "Britain’s national curry, sattvic version. Smoked paprika stands in for tandoor and alliums.",
          ["2 tbsp ghee or butter", "1 bay leaf", "4 cardamom", "Pinch of hing",
           "3 tomatoes + 12 soaked cashews, blended", "1 tsp ginger paste",
           "1 tsp Kashmiri chilli", "1 tsp coriander powder", "½ tsp smoked paprika",
           "¼ tsp turmeric", "Cauliflower florets, capsicum, paneer cubes, peas (about 4 cups mixed, roasted)",
           "½ cup water", "3 tbsp cream", "1 tsp kasuri methi", "½ tsp garam masala", "Salt", "Pinch of sugar"],
          ["Roast the vegetables and paneer on a steel tray until charred at the edges.",
           "Make the gravy: ghee, whole spices, hing, tomato-cashew paste and ground spices. Cook until fat separates.",
           "Add water, cream, kasuri methi, salt and sugar.",
           "Fold in roasted vegetables. Rest 5 minutes before serving."]),
        r("europe", "Side", "Dal Tadka", "British Indian / North Indian", 10, 30, CLAY,
          "The side that makes tikka masala a proper meal.",
          ["½ cup toor dal", "½ cup moong dal", "¼ tsp turmeric", "Salt", "3 cups water",
           "1 tomato, chopped", "2 tbsp ghee", "1 tsp cumin seeds", "Pinch of hing",
           "2 green chillies, slit", "½ tsp coriander powder", "¼ tsp chilli powder", "Lemon", "Coriander"],
          ["Pressure-cook or simmer dals with turmeric until soft. Whisk smooth.",
           "Stir in tomato and salt. Simmer 5 minutes.",
           "Tadka: hot ghee, cumin, hing, chilli. Pour over the dal. Add remaining spices, lemon and coriander."],
          "Clay pot after cooking gives a rounder flavour. Never store leftover dal in aluminium."),
        r("europe", "Bread", "Butter Naan (no garlic)", "United Kingdom", 70, 15, TAWA,
          "Restaurant naan without garlic butter.",
          ["2 cups maida", "½ cup yogurt", "½ tsp sugar", "½ tsp salt", "½ tsp baking powder",
           "2 tbsp milk", "1 tbsp oil", "Butter and coriander for finishing"],
          ["Knead a soft dough. Rest 1 hour.",
           "Roll ovals. Cook on a screaming-hot iron tawa. Flip and blister the top over a flame.",
           "Brush with butter (not garlic butter) and coriander."],
          "If you have no flame, finish 2 minutes under a hot steel-grill oven."),
        r("europe", "Sweet", "Kalakand", "British Indian mithai counter", 10, 25, S,
          "Milk-cake sweet sold in every UK Indian sweet shop. Naturally no allium.",
          ["500 g full-fat ricotta or well-drained chenna", "200 g condensed milk",
           "1 tsp ghee", "½ tsp cardamom", "Pistachios"],
          ["Cook ricotta and condensed milk in a steel kadhai on medium-low, stirring, 15–20 minutes until thick.",
           "Add ghee and cardamom. Spread in a greased steel/glass tray.",
           "Top with pistachios. Chill 1 hour and cut."]),
        r("europe", "Dessert", "Kesar Shrikhand", "UK / Maharashtrian", 10, 0, "Steel or glass bowl — no cooking pan needed",
          "No-cook yogurt dessert. The calm end to a masala-heavy British Indian meal.",
          ["500 g hung yogurt", "⅓ cup powdered sugar", "8 saffron strands in 1 tbsp warm milk",
           "½ tsp cardamom", "Pistachios and almonds"],
          ["Whisk hung yogurt with sugar, saffron milk and cardamom until satin-smooth.",
           "Chill 2 hours. Serve with nuts.",
           "Do not add fruit with onion-family garnishes (spring onion is sometimes used on fusion plates — skip it)."]),
        r("europe", "Salad", "No-Onion Kachumber", "British Indian table salad", 10, 0, "Glass bowl",
          "Pub-curry-house salad minus the raw onion.",
          ["1 cucumber, diced", "2 tomatoes, diced", "½ green apple or raw mango, diced (for bite)",
           "1 green chilli", "Coriander", "Lemon", "Roasted cumin", "Salt", "Pinch of sugar"],
          ["Combine and dress just before serving.",
           "Apple or raw mango replaces the crunch and bite of onion."]),

        # ---------- NORTH AMERICA ----------
        r("north-america", "Starter", "Aloo-Pea Samosa (no onion)", "Punjabi-Canadian", 40, 25, "Steel kadhai for frying + steel bowl for filling",
          "The starter of every North American Indian party. Filling is potato, pea and spice — no onion.",
          ["2 cups maida", "4 tbsp oil (for dough)", "Salt", "Water",
           "3 boiled potatoes, mashed", "½ cup green peas", "1 tsp cumin seeds", "Pinch of hing",
           "1 tsp ginger-green chilli paste", "½ tsp turmeric", "1 tsp coriander powder",
           "½ tsp garam masala", "½ tsp amchur", "Oil for frying"],
          ["Rub oil into maida and salt. Knead a firm dough. Rest 20 minutes.",
           "Filling: heat 1 tbsp oil, cumin and hing. Add ginger-chilli, peas, potato and spices. Cool.",
           "Roll dough into ovals, cut in half, form cones, fill, seal with water.",
           "Fry in medium-hot oil in a steel kadhai until golden. Serve with tamarind and mint chutney (mint chutney without onion/garlic)."]),
        r("north-america", "Main", "Restaurant Chana Masala", "Punjabi-Canadian / US", 15, 25, S,
          "The vegetarian main of Indian North America.",
          ["2 cups cooked chickpeas (or 1 can, rinsed)", "2 tbsp oil", "1 tsp cumin seeds", "Pinch of hing",
           "2 tomatoes blended with 1-inch ginger", "1 tsp chana masala powder (check: no garlic)",
           "1 tsp coriander powder", "½ tsp chilli powder", "¼ tsp turmeric", "½ tsp amchur",
           "Salt", "Lemon", "Julienned ginger", "Coriander"],
          ["Temper cumin and hing. Add tomato-ginger paste and dry spices. Cook until thick and glossy.",
           "Add chickpeas and a splash of water. Simmer 12 minutes. Mash a few for body.",
           "Finish with amchur, lemon, ginger strips and coriander."],
          "Many North American 'curry pastes' are garlic-heavy — make the gravy from tomatoes instead."),
        r("north-america", "Side", "Palak Paneer (no onion garlic)", "Punjabi North America", 15, 20, S,
          "Saag paneer without the usual onion-garlic fry-up.",
          ["300 g spinach, washed", "200 g paneer", "1 tbsp ghee + 1 tbsp oil",
           "1 tsp cumin", "Pinch of hing", "1 tsp ginger paste", "1 green chilli",
           "1 small tomato", "½ tsp coriander powder", "¼ tsp garam masala",
           "2 tbsp cream or yogurt", "Salt", "Pinch of sugar"],
          ["Blanch spinach 1 minute, plunge in cold water, blend with chilli and tomato.",
           "Heat ghee and oil. Cumin, hing, ginger. Add spinach puree and spices. Cook 6–8 minutes.",
           "Add paneer, salt, sugar and cream. Do not boil hard after the cream goes in."]),
        r("north-america", "Bread", "Potato Kulcha (no onion)", "Punjabi-Canadian", 40, 20, TAWA,
          "Amritsari-style stuffed bread minus chopped onion in the filling.",
          ["2 cups maida", "½ cup yogurt", "½ tsp sugar", "½ tsp salt", "½ tsp baking powder", "2 tbsp oil",
           "2 boiled potatoes", "1 tsp ginger-chilli paste", "½ tsp amchur", "½ tsp cumin",
           "Coriander", "Butter"],
          ["Knead dough with yogurt, sugar, salt, baking powder and oil. Rest 30 minutes.",
           "Mash potato with ginger-chilli, amchur, cumin, salt and coriander — no onion.",
           "Stuff, roll gently, cook on a hot iron tawa with butter until brown spots appear."]),
        r("north-america", "Sweet", "Besan Barfi", "US / Canada Indian mithai", 10, 25, S,
          "Gram-flour fudge. The sweet box staple of North American Indian homes.",
          ["1 cup ghee", "2 cups besan", "1¼ cups sugar", "⅓ cup water", "½ tsp cardamom", "Pistachios"],
          ["Roast besan in ghee on low 12–15 minutes until nutty and the ghee surfaces. Steel kadhai only.",
           "Make one-thread sugar syrup with sugar and water.",
           "Mix syrup into the besan off the heat. Add cardamom. Spread in a steel/glass tray. Set and cut."]),
        r("north-america", "Dessert", "Cardamom Rice Pudding Kheer", "US Indian home cooking", 10, 40, "Heavy stainless steel pot",
          "Same as Indian kheer, often finished with maple or extra nuts in North America. Keep it classic here.",
          ["¼ cup basmati", "1 litre milk", "⅓ cup sugar", "Cardamom", "Saffron optional", "Almonds", "Raisins"],
          ["Simmer rice in milk in a steel pot 35–40 minutes, stirring.",
           "Add sugar, cardamom and nuts. Serve warm or cold."],
          "Do not cook this in a slow-cooker with an aluminium insert."),
        r("north-america", "Salad", "Sprouted Moong Chaat", "US / Canada", 15, 0, "Steel or glass bowl",
          "Protein salad that replaced onion-heavy bhel at many North American vegetarian tables.",
          ["2 cups sprouted moong", "1 tomato, diced", "½ cup pomegranate", "Coriander",
           "Lemon", "Chaat masala", "Black salt", "1 green chilli", "Optional: diced cucumber"],
          ["Toss just before eating. Taste for lemon and black salt.",
           "No onion. Pomegranate gives the pop."]),

        # ---------- SOUTH AMERICA / CARIBBEAN ----------
        r("south-america", "Starter", "Pholourie", "Trinidad", 20, 20, "Steel kadhai for frying",
          "Split-pea fritters sold with doubles. Garlic is usual — hing and geera replace it here.",
          ["1 cup split pea flour (or ground soaked yellow split peas)", "½ cup flour",
           "1 tsp baking powder", "1 tsp geera (cumin), toasted and ground", "Pinch of hing",
           "1 tsp ginger", "½ tsp turmeric", "1 green chilli, minced", "Salt", "Water", "Oil for frying"],
          ["Mix to a thick dropping batter. Rest 15 minutes.",
           "Drop spoonfuls into medium oil. Fry until golden and cooked through.",
           "Serve with tamarind sauce (no garlic) or mango kuchela without onion."]),
        r("south-america", "Main", "Channa and Aloo", "Trinidad and Guyana", 15, 30, S,
          "The soul of doubles and veg roti. Curry powder plus geera, not onion and garlic.",
          ["2 tbsp oil", "1 tsp geera (cumin seeds)", "Pinch of hing",
           "1 tbsp curry powder (label must not list garlic or onion)",
           "1 boiled potato, cubed", "1½ cups cooked channa (chickpeas)",
           "1 tomato, chopped", "½ tsp turmeric", "1 tsp coriander powder",
           "1 whole Scotch bonnet (optional, do not burst)", "Water to cover", "Salt",
           "Culantro (bandhania) or coriander", "Lime"],
          ["Heat oil. Bloom geera and hing. Add curry powder and fry 30 seconds.",
           "Add tomato, turmeric, coriander, potato, channa, salt and water.",
           "Tuck in the whole Scotch bonnet. Simmer 15–20 minutes until thick.",
           "Remove the pepper. Finish with culantro and lime."],
          "Caribbean curry powder often hides garlic. Use a garlic-free blend or mix coriander, cumin, turmeric and fenugreek yourself."),
        r("south-america", "Side", "Baigan Choka (no garlic)", "Trinidad", 10, 25, "Cast-iron or open flame + steel bowl",
          "Smoky mashed eggplant. Classic choka is heavy on garlic — hing, cumin and hot pepper do that job.",
          ["2 large baigan (eggplant)", "2 tbsp oil", "½ tsp toasted geera", "Pinch of hing",
           "1 tomato, roasted", "Hot pepper to taste", "Salt", "Culantro or coriander"],
          ["Roast whole baigan on a flame or in a cast-iron pan until collapsed and charred. Peel.",
           "Mash with roasted tomato, oil, geera, hing, pepper and salt.",
           "Taste. If it feels flat, add more toasted geera and lime — not garlic."]),
        r("south-america", "Bread", "Sada Roti", "Trinidad / Guyana", 30, 15, TAWA,
          "Thick Indo-Caribbean bread for scooping channa. Iron tawa only.",
          ["3 cups atta", "Pinch of baking powder", "Pinch of salt", "Water"],
          ["Knead a soft dough. Rest 20 minutes.",
           "Pat into a thick round. Cook on a hot iron tawa, turning, until it puffs and spots.",
           "Wrap in a cloth. Tear and scoop the curry."]),
        r("south-america", "Sweet", "Parsad", "Trinidad Hindu prayer sweet", 10, 20, S,
          "Cream-of-wheat prasad. Already onion-garlic free; keep it in steel.",
          ["½ cup ghee", "1 cup sooji (cream of wheat)", "½ cup sugar", "2 cups milk",
           "4 cardamom", "2 tbsp raisins", "2 tbsp chopped cherries or mixed peel (optional)"],
          ["Roast sooji in ghee on low until fragrant.",
           "Warm milk with sugar and cardamom separately in steel.",
           "Pour into the sooji, stirring. Add raisins. Cook until it leaves the sides.",
           "Offer, then serve warm."]),
        r("south-america", "Dessert", "Caribbean Sweet Rice", "Guyana / Trinidad", 10, 40, "Heavy stainless steel pot",
          "Indo-Caribbean kheer, often with coconut milk on the islands.",
          ["¼ cup rice", "2 cups milk", "1 cup coconut milk", "⅓ cup sugar",
           "Cardamom", "Cinnamon stick", "Raisins", "Vanilla (optional, ¼ tsp)"],
          ["Simmer rice in milk and cinnamon in a steel pot until very soft.",
           "Add coconut milk, sugar, cardamom and raisins. Cook until creamy.",
           "Serve chilled or warm. No aluminium pot."]),
        r("south-america", "Salad", "Green Mango Salad", "Trinidad", 15, 0, "Glass bowl",
          "Sharp salad that replaces onion kuchela on a sattvic table.",
          ["2 green mangoes, julienned", "1 cucumber, julienned", "Hot pepper, minced",
           "Lime", "Salt", "Roasted geera", "Coriander or culantro", "Pinch of sugar"],
          ["Toss mango and cucumber with lime, salt, geera and pepper.",
           "Rest 10 minutes. The mango’s acid is the ‘onion’."]),

        # ---------- OCEANIA ----------
        r("oceania", "Starter", "Vegetable Pakora", "Indo-Fijian", 15, 15, "Steel kadhai for frying",
          "Girmitiya snack. Mixed veg, no onion — cabbage, spinach and potato bind with besan.",
          ["1 cup gram flour", "2 tbsp rice flour", "1 cup shredded cabbage", "½ cup chopped spinach",
           "1 small potato, grated", "1 tsp ginger-chilli paste", "½ tsp turmeric", "1 tsp cumin",
           "Pinch of hing", "Salt", "Water", "Oil"],
          ["Mix vegetables with salt and spices, then flours and a little water.",
           "Fry medium-size pakoras in a steel kadhai until crisp.",
           "Serve with tamarind chutney."]),
        r("oceania", "Main", "Jackfruit Khatar Curry", "Indo-Fijian", 15, 30, S,
          "Fiji’s vegetarian jackfruit curry, originally onion-garlic heavy. Hing, curry leaves and ginger take over.",
          ["400 g young jackfruit (fresh or tinned, drained)", "3 tbsp oil",
           "½ tsp mustard seeds", "½ tsp cumin seeds", "10 curry leaves", "Pinch of hing",
           "1 tsp ginger paste", "½ tsp turmeric", "2 tsp coriander powder", "1 tsp chilli powder",
           "1 tsp garlic-free curry powder", "½ cup water", "Salt", "Coriander"],
          ["Heat oil. Splutter mustard, cumin, curry leaves and hing.",
           "Add ginger and dry spices. Fry 30 seconds.",
           "Add jackfruit, salt and water. Cook 15–20 minutes until almost dry and the edges catch colour.",
           "Garnish with coriander."],
          "Tinned jackfruit in brine: rinse well. Tinned in syrup is for dessert, not this curry."),
        r("oceania", "Side", "Fijian Dhal", "Indo-Fijian", 10, 35, CLAY,
          "The everyday soup-dal of Fiji Indian homes.",
          ["1 cup masoor or toor dal", "¼ tsp turmeric", "4 cups water", "Salt",
           "2 tbsp oil", "1 tsp mustard seeds", "1 tsp cumin", "Pinch of hing", "8 curry leaves",
           "2 dry chillies", "1 tomato", "Lemon"],
          ["Cook dal with turmeric until very soft. Whisk.",
           "Add tomato and salt. Simmer 5 minutes.",
           "Tadka: oil, mustard, cumin, hing, curry leaves, chilli. Pour over. Lemon at the table."]),
        r("oceania", "Bread", "Fijian Roti", "Indo-Fijian", 25, 15, TAWA,
          "Soft roti eaten with khatar and dhal.",
          ["2 cups atta", "1 tbsp oil", "Pinch of salt", "Warm water", "Ghee"],
          ["Knead soft dough. Rest 15 minutes. Roll and cook on iron tawa. Ghee between layers of the stack."]),
        r("oceania", "Sweet", "Milk-Powder Gulab Jamun", "Indo-Fijian / wider Indian", 30, 25, "Steel kadhai + steel saucepan for syrup",
          "Festival sweet across Fiji. Already no allium. Syrup must not touch aluminium.",
          ["1 cup milk powder", "¼ cup maida", "¼ tsp baking soda", "2 tbsp ghee", "Milk to bind",
           "1½ cups sugar", "1½ cups water", "4 cardamom", "Few saffron strands", "Oil or ghee for frying"],
          ["Make a one-thread-adjacent syrup: sugar, water, cardamom, saffron. Keep warm in steel.",
           "Rub ghee into milk powder, maida and soda. Bind with milk to a soft dough. Rest 10 minutes.",
           "Shape smooth balls. Fry on low-medium until deep brown.",
           "Drain and soak in warm syrup at least 30 minutes."],
          "If the syrup is made in aluminium it can crystallise oddly and taste sharp."),
        r("oceania", "Dessert", "Almond Phirni", "Indo-Fijian festive table", 20, 25, S,
          "Ground-rice milk dessert set in steel or clay pots.",
          ["¼ cup basmati, soaked and ground to a paste", "3 cups milk", "⅓ cup sugar",
           "Saffron", "Cardamom", "Slivered almonds"],
          ["Boil milk in steel. Stir in rice paste. Cook until it thickens and the raw taste goes.",
           "Add sugar, saffron and cardamom. Pour into steel katoris. Chill. Top with almonds."]),
        r("oceania", "Salad", "Cucumber Tomato Coriander Salad", "Fiji", 10, 0, "Glass bowl",
          "Simple island table salad, no onion.",
          ["1 cucumber", "2 tomatoes", "1 green chilli", "Lemon", "Salt", "Coriander", "Roasted cumin"],
          ["Dice, dress, serve at once beside the dhal and roti."]),

        # ---------- ANTARCTICA ----------
        r("antarctica", "Starter", "Masala Makhana", "Polar / expedition snack", 5, 10, S,
          "Fox-nut snack from dry stores. No fresh onion, no aluminium wok needed — steel kadhai on a camp stove.",
          ["4 cups makhana (fox nuts)", "1 tbsp ghee", "Pinch of hing", "½ tsp turmeric",
           "½ tsp chilli powder", "½ tsp chaat masala", "Salt"],
          ["Warm ghee. Add hing, then makhana. Roast on medium-low until crisp (5–8 minutes).",
           "Dust with turmeric, chilli, chaat masala and salt. Cool — they crisp further.",
           "Store in a steel or glass tin, not foil."]),
        r("antarctica", "Main", "Tomato-Cashew Dal with Dried Vegetables", "Polar kitchen", 10, 25, S,
          "Shelf-stable main: dal + tomato puree + cashew powder + freeze-dried mixed veg. Tastes like shahi dal.",
          ["¾ cup toor or moong dal (dry)", "3 tbsp tomato puree (tin or tube)", "2 tbsp cashew powder",
           "1 cup freeze-dried mixed vegetables, rehydrated", "2 tbsp ghee", "1 tsp cumin",
           "Pinch of hing", "1 tsp ginger powder or 1 tsp ginger paste (tube)",
           "1 tsp Kashmiri chilli", "1 tsp coriander powder", "¼ tsp turmeric",
           "Salt", "1 tbsp milk powder + a little water (for creaminess)"],
          ["Cook dal in a steel pot until soft (pressure cooker if the station has one).",
           "Tadka: ghee, cumin, hing, ginger, tomato puree and spices. Cook until thick.",
           "Add dal, rehydrated veg, cashew powder and milk-powder cream. Simmer 5 minutes."],
          "Do not cook tomato puree in aluminium mess tins — use the station’s stainless insert."),
        r("antarctica", "Side", "Canned Chickpea Tadka", "Polar kitchen", 5, 12, S,
          "Emergency chana from cans. Drain, rinse, tadka.",
          ["1 can chickpeas, rinsed", "2 tbsp oil", "1 tsp cumin", "Pinch of hing",
           "2 tbsp tomato puree", "½ tsp turmeric", "1 tsp coriander", "½ tsp chilli",
           "½ tsp amchur or lemon crystals", "Salt"],
          ["Bloom cumin and hing in oil. Add puree and spices. Cook 2 minutes.",
           "Add chickpeas. Simmer 8 minutes. Finish with amchur."]),
        r("antarctica", "Bread", "Station Atta Roti", "Polar kitchen", 20, 15, TAWA,
          "Flour, water, salt, iron skillet. Works on a research-station range.",
          ["2 cups atta (packed dry)", "Warm water", "Pinch of salt", "Ghee if available"],
          ["Knead, rest 10 minutes if you can, roll, cook on a cast-iron skillet. Keep stacked in a cloth.",
           "No aluminium griddle — it warps and cooks pale rotis."]),
        r("antarctica", "Sweet", "Condensed-Milk Coconut Ladoo", "Polar kitchen", 5, 10, S,
          "Desiccated coconut + condensed milk. Both survive the journey south.",
          ["2 cups desiccated coconut", "¾ cup condensed milk", "½ tsp cardamom powder", "1 tsp ghee"],
          ["Cook coconut, condensed milk and ghee in a steel pan until thick.",
           "Add cardamom. Shape. Roll in extra coconut."]),
        r("antarctica", "Dessert", "Milk-Powder Saffron Kheer", "Polar kitchen", 5, 15, S,
          "No litres of fresh milk. Milk powder plus a handful of rice (or broken vermicelli).",
          ["½ cup milk powder mixed with 3 cups water (or as pack directs for whole milk)",
           "3 tbsp rice or roasted vermicelli", "3 tbsp sugar", "Saffron or cardamom from the spice kit",
           "Dried raisins and nuts"],
          ["Bring reconstituted milk to a simmer in steel. Add rice or vermicelli; cook until soft.",
           "Add sugar, saffron/cardamom and dried fruit. Serve in steel mugs."],
          "Do not reduce reconstituted milk in aluminium — it scorches and tastes of metal faster than fresh milk."),
        r("antarctica", "Salad", "Pickled Carrot and Cucumber Jar Salad", "Polar kitchen", 15, 0, "Glass jar (never aluminium tin for storage)",
          "Fresh salad is scarce. Quick-pickle dry-store carrot and any cucumber from hydroponics, or use jarred gherkin plus carrot.",
          ["2 carrots, julienned (fresh, hydroponic, or rehydrated strips)",
           "1 cucumber or 4 crunchy pickles, sliced", "1 tsp salt", "1 tsp sugar",
           "3 tbsp lemon juice or vinegar", "½ tsp roasted cumin", "Pinch of chilli flakes", "Coriander if available"],
          ["Toss with salt, sugar and acid. Rest 20 minutes (or keep in a glass jar 1 day).",
           "Store only in glass or steel. Acid plus aluminium jars is unsafe."],
          "This is the polar stand-in for kachumber when onion is both forbidden and unavailable."),
    ]


# Override this list when generating another region set (e.g. Indian states).
REGIONS: list[dict] | None = None


def continent_by_id(cid: str) -> dict:
    pool = REGIONS if REGIONS is not None else CONTINENTS
    return next(c for c in pool if c["id"] == cid)


def md_for(recipe: dict, continent: dict) -> str:
    ing = "\n".join(f"- {i}" for i in recipe["ingredients"])
    steps = "\n".join(f"{n}. {s}" for n, s in enumerate(recipe["method"], 1))
    notes = f"\n## Notes\n\n{recipe['notes']}\n" if recipe["notes"] else ""
    return f"""# {recipe['name']}

**Continent:** {continent['name']} — {recipe['community']}  
**Category:** {recipe['category']}  
**Diet:** {recipe['diet']}  
**Servings:** {recipe['servings']}  
**Prep:** {recipe['prep_min']} min  
**Cook:** {recipe['cook_min']} min  
**Cookware:** {recipe['cookware']}

## Why this dish

{recipe['why']}

## Ingredients

{ing}

## Method

{steps}
{notes}"""


def style_header_row(ws, row, cols, fill=SAFFRON):
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN


def banner(ws, title, subtitle, cols=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(1, 1, title)
    c.fill = SAFFRON
    c.font = TITLE_FONT
    c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    s = ws.cell(2, 1, subtitle)
    s.fill = GREEN
    s.font = SUB_FONT
    s.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 22
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=cols)
    w = ws.cell(
        3,
        1,
        "VEGETARIAN  ·  NO ONION  ·  NO GARLIC  ·  NO ALLIUM  ·  NO ALUMINIUM COOKWARE",
    )
    w.fill = RED
    w.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    w.alignment = CENTER
    ws.row_dimensions[3].height = 20
    for col in range(1, cols + 1):
        ws.cell(1, col).fill = SAFFRON
        ws.cell(2, col).fill = GREEN
        ws.cell(3, col).fill = RED


def set_widths(ws, widths: dict[int, int]):
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def write_recipe_sheet(ws, recipe: dict, continent: dict):
    banner(
        ws,
        recipe["name"],
        f"{continent['name']}  ·  {recipe['category']}  ·  {recipe['community']}",
        cols=4,
    )
    meta = [
        ("Servings", recipe["servings"]),
        ("Prep (min)", recipe["prep_min"]),
        ("Cook (min)", recipe["cook_min"]),
        ("Cookware", recipe["cookware"]),
        ("Diet", recipe["diet"]),
        ("Why this dish", recipe["why"]),
    ]
    row = 5
    ws.cell(row, 1, "Field").font = HEADER_FONT
    ws.cell(row, 1).fill = SLATE
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    ws.cell(row, 2, "Detail").font = HEADER_FONT
    ws.cell(row, 2).fill = SLATE
    for col in range(1, 5):
        ws.cell(row, col).fill = SLATE
        ws.cell(row, col).font = HEADER_FONT
        ws.cell(row, col).border = THIN
    row = 6
    for label, value in meta:
        ws.cell(row, 1, label).font = BOLD
        ws.cell(row, 1).fill = SAND
        ws.cell(row, 1).border = THIN
        ws.cell(row, 1).alignment = WRAP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        cell = ws.cell(row, 2, value)
        cell.alignment = WRAP
        cell.font = BODY_FONT
        cell.border = THIN
        for col in range(2, 5):
            ws.cell(row, col).border = THIN
            ws.cell(row, col).fill = CREAM
        ws.row_dimensions[row].height = 36 if label in ("Cookware", "Why this dish", "Diet") else 22
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    h = ws.cell(row, 1, "Ingredients")
    h.fill = GOLD
    h.font = HEADER_FONT
    for col in range(1, 5):
        ws.cell(row, col).fill = GOLD
    row += 1
    ws.cell(row, 1, "#").fill = SLATE
    ws.cell(row, 1).font = HEADER_FONT
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    ws.cell(row, 2, "Item").font = HEADER_FONT
    ws.cell(row, 2).fill = SLATE
    for col in range(1, 5):
        ws.cell(row, col).fill = SLATE
        ws.cell(row, col).font = HEADER_FONT
        ws.cell(row, col).border = THIN
    row += 1
    for i, item in enumerate(recipe["ingredients"], 1):
        ws.cell(row, 1, i).alignment = CENTER
        ws.cell(row, 1).border = THIN
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row, 2, item).alignment = WRAP
        ws.cell(row, 2).border = THIN
        fill = CREAM if i % 2 else WHITE
        for col in range(1, 5):
            ws.cell(row, col).fill = fill
            ws.cell(row, col).border = THIN
            ws.cell(row, col).font = BODY_FONT
        ws.row_dimensions[row].height = 22
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    h = ws.cell(row, 1, "Method")
    h.fill = GREEN
    h.font = HEADER_FONT
    for col in range(1, 5):
        ws.cell(row, col).fill = GREEN
    row += 1
    ws.cell(row, 1, "Step").fill = SLATE
    ws.cell(row, 1).font = HEADER_FONT
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    ws.cell(row, 2, "What to do").font = HEADER_FONT
    ws.cell(row, 2).fill = SLATE
    for col in range(1, 5):
        ws.cell(row, col).fill = SLATE
        ws.cell(row, col).font = HEADER_FONT
        ws.cell(row, col).border = THIN
    row += 1
    for i, step in enumerate(recipe["method"], 1):
        ws.cell(row, 1, i).alignment = CENTER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row, 2, step).alignment = WRAP
        fill = GREEN_SOFT if i % 2 else WHITE
        for col in range(1, 5):
            ws.cell(row, col).fill = fill
            ws.cell(row, col).border = THIN
            ws.cell(row, col).font = BODY_FONT
        ws.row_dimensions[row].height = 48
        row += 1

    if recipe["notes"]:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        n = ws.cell(row, 1, "Notes")
        n.fill = RED
        n.font = HEADER_FONT
        for col in range(1, 5):
            ws.cell(row, col).fill = RED
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        note = ws.cell(row, 1, recipe["notes"])
        note.alignment = WRAP
        note.fill = RED_SOFT
        note.font = BODY_FONT
        for col in range(1, 5):
            ws.cell(row, col).fill = RED_SOFT
            ws.cell(row, col).border = THIN
        ws.row_dimensions[row].height = 48

    set_widths(ws, {1: 22, 2: 36, 3: 22, 4: 36})
    ws.freeze_panes = "A5"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_title_rows = "1:3"
    ws.sheet_properties.pageSetUpPr.fitToPage = True


HEADERS = [
    "Continent",
    "Community",
    "Category",
    "Recipe",
    "Servings",
    "Prep min",
    "Cook min",
    "Cookware",
    "Ingredients",
    "Method",
    "Notes",
    "Diet",
    "Folder path",
]


def recipe_row(recipe: dict, continent: dict) -> list:
    base = continent.get(
        "path_prefix",
        f"recipes/onion-garlic-free-indian/{continent['folder']}",
    )
    folder = f"{base}/{CAT_FOLDERS[recipe['category']]}/{slug(recipe['name'])}.md"
    return [
        continent["name"],
        recipe["community"],
        recipe["category"],
        recipe["name"],
        recipe["servings"],
        recipe["prep_min"],
        recipe["cook_min"],
        recipe["cookware"],
        "\n".join(f"• {i}" for i in recipe["ingredients"]),
        "\n".join(f"{n}. {s}" for n, s in enumerate(recipe["method"], 1)),
        recipe["notes"],
        recipe["diet"],
        folder,
    ]


def write_table_sheet(ws, title, subtitle, rows: list[list], freeze="A5"):
    cols = len(HEADERS)
    banner(ws, title, subtitle, cols=cols)
    header_row = 5
    for i, h in enumerate(HEADERS, 1):
        cell = ws.cell(header_row, i, h)
        cell.fill = SLATE
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN
    for r_i, data in enumerate(rows, 6):
        cat = data[2]
        for c_i, value in enumerate(data, 1):
            cell = ws.cell(r_i, c_i, value)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN
            if c_i == 3:
                cell.fill = CATEGORY_FILL.get(str(cat), CREAM)
                cell.alignment = CENTER
                cell.font = BOLD
            elif r_i % 2 == 0:
                cell.fill = CREAM
            else:
                cell.fill = WHITE
        ws.row_dimensions[r_i].height = 90
    set_widths(
        ws,
        {
            1: 16,
            2: 28,
            3: 12,
            4: 32,
            5: 11,
            6: 11,
            7: 11,
            8: 34,
            9: 46,
            10: 52,
            11: 36,
            12: 28,
            13: 42,
        },
    )
    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A5:{get_column_letter(cols)}{5 + len(rows)}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:5"


def write_menu_sheet(ws, continent: dict, recipes: list[dict]):
    banner(
        ws,
        f"{continent['name']} — onion-garlic-free vegetarian Indian menu",
        continent["community"],
        cols=7,
    )
    ws.merge_cells("A5:G5")
    n = ws.cell(5, 1, continent["note"])
    n.alignment = WRAP
    n.fill = SAND
    n.font = BODY_FONT
    for col in range(1, 8):
        ws.cell(5, col).fill = SAND
        ws.cell(5, col).border = THIN
    ws.row_dimensions[5].height = 40

    headers = ["Course", "Recipe", "Prep min", "Cook min", "Servings", "Cookware", "Markdown file"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(7, i, h)
        cell.fill = SLATE
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN
    order = ["Starter", "Main", "Side", "Bread", "Sweet", "Dessert", "Salad"]
    by_cat = {c: [] for c in order}
    for rec in recipes:
        by_cat[rec["category"]].append(rec)
    row = 8
    for cat in order:
        for rec in by_cat[cat]:
            path = f"{CAT_FOLDERS[cat]}/{slug(rec['name'])}.md"
            values = [cat, rec["name"], rec["prep_min"], rec["cook_min"], rec["servings"], rec["cookware"], path]
            for i, v in enumerate(values, 1):
                cell = ws.cell(row, i, v)
                cell.font = BODY_FONT
                cell.alignment = WRAP
                cell.border = THIN
                if i == 1:
                    cell.fill = CATEGORY_FILL[cat]
                    cell.font = BOLD
                    cell.alignment = CENTER
                elif row % 2 == 0:
                    cell.fill = CREAM
            ws.row_dimensions[row].height = 32
            row += 1
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=7)
    foot = ws.cell(
        row + 1,
        1,
        "Cook in stainless steel, cast iron, clay, enamel or glass. Never aluminium. No onion, garlic, shallot, leek or chive.",
    )
    foot.fill = RED_SOFT
    foot.font = BOLD
    foot.alignment = WRAP
    for col in range(1, 8):
        ws.cell(row + 1, col).fill = RED_SOFT
    ws.row_dimensions[row + 1].height = 28
    set_widths(ws, {1: 14, 2: 36, 3: 12, 4: 12, 5: 12, 6: 42, 7: 40})
    ws.freeze_panes = "A8"


def write_shopping_sheet(ws, recipes: list[dict], title: str):
    banner(ws, title, "Tick off what you already have. All items are onion-garlic free.", cols=5)
    headers = ["Ingredient", "Recipe", "Continent", "Category", "Got it?"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(5, i, h)
        cell.fill = SLATE
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN
    row = 6
    for rec in recipes:
        continent = continent_by_id(rec["continent_id"])
        for item in rec["ingredients"]:
            ws.cell(row, 1, item).alignment = WRAP
            ws.cell(row, 2, rec["name"])
            ws.cell(row, 3, continent["name"])
            ws.cell(row, 4, rec["category"])
            ws.cell(row, 5, "☐")
            ws.cell(row, 5).alignment = CENTER
            fill = CATEGORY_FILL.get(rec["category"], CREAM)
            ws.cell(row, 4).fill = fill
            for col in range(1, 6):
                ws.cell(row, col).border = THIN
                ws.cell(row, col).font = BODY_FONT
                if col != 4:
                    ws.cell(row, col).fill = CREAM if row % 2 == 0 else WHITE
            ws.row_dimensions[row].height = 20
            row += 1
    dv = DataValidation(type="list", formula1='"☐,☑"', allow_blank=True)
    dv.error = "Use the tick box"
    dv.errorTitle = "Got it?"
    ws.add_data_validation(dv)
    dv.add(f"E6:E{row - 1}")
    ws.auto_filter.ref = f"A5:E{row - 1}"
    ws.freeze_panes = "A6"
    set_widths(ws, {1: 55, 2: 32, 3: 18, 4: 12, 5: 12})
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def write_rules_sheet(ws):
    banner(ws, "Diet and cookware rules", "Apply these to every recipe in this workbook", cols=4)
    blocks = [
        ("Diet", [
            "Vegetarian — no meat, fish or eggs.",
            "No onion, garlic, shallot, leek, chive or spring onion.",
            "No onion powder or garlic powder. Read curry-powder labels.",
            "Ginger, hing (asafoetida), tomato, coconut, yogurt and whole spices are allowed.",
            "Hing in hot ghee or oil replaces onion and garlic. Use a pinch only.",
        ]),
        ("Cookware — never aluminium", [
            "Use stainless steel pots, kadhai, steamers and thalis.",
            "Use cast-iron tawa for breads.",
            "Clay, enamel, glass or ceramic are fine.",
            "Do not use aluminium pots, steamers, foil trays or aluminium pressure-cooker inners for tomato gravies.",
            "Tomato, tamarind, lemon, yogurt and milk all react with aluminium.",
        ]),
        ("Instead of onion and garlic", [
            "Savoury depth → pinch of hing in hot fat.",
            "Fresh heat → ginger + green chilli.",
            "Gravy body → tomato puree + cashew paste.",
            "Restaurant finish → kasuri methi + garam masala + cream.",
            "South Indian snap → mustard seeds, urad dal, curry leaves, coconut.",
            "Tang → lemon, amchur, tamarind, yogurt.",
        ]),
    ]
    row = 5
    for title, lines in blocks:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        h = ws.cell(row, 1, title)
        heading_fill = RED if "aluminium" in title.lower() else GOLD
        h.fill = heading_fill
        h.font = HEADER_FONT
        for col in range(1, 5):
            ws.cell(row, col).fill = heading_fill
        row += 1
        for line in lines:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            cell = ws.cell(row, 1, "• " + line)
            cell.alignment = WRAP
            cell.font = BODY_FONT
            fill = RED_SOFT if "aluminium" in title.lower() else CREAM
            for col in range(1, 5):
                ws.cell(row, col).fill = fill
                ws.cell(row, col).border = THIN
            ws.row_dimensions[row].height = 22
            row += 1
        row += 1
    set_widths(ws, {1: 28, 2: 28, 3: 28, 4: 28})


def write_cover_sheet(ws, recipes: list[dict]):
    banner(
        ws,
        "All continents — vegetarian Indian feast",
        "Onion-garlic-free  ·  no aluminium  ·  starter to dessert",
        cols=6,
    )
    ws.merge_cells("A5:F5")
    intro = ws.cell(
        5,
        1,
        "One complete vegetarian Indian menu per continent. Open each continent sheet, or use All Recipes with filters. "
        "Each continent folder on disk also has markdown files and its own Excel workbook.",
    )
    intro.alignment = WRAP
    intro.fill = SAND
    for col in range(1, 7):
        ws.cell(5, col).fill = SAND
    ws.row_dimensions[5].height = 40

    headers = ["Continent", "Community", "Recipes", "Folder", "Workbook", "Courses"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(7, i, h)
        cell.fill = SLATE
        cell.font = HEADER_FONT
        cell.border = THIN
        cell.alignment = CENTER
    row = 8
    for c in CONTINENTS:
        recs = [x for x in recipes if x["continent_id"] == c["id"]]
        values = [
            c["name"],
            c["community"],
            len(recs),
            f"recipes/onion-garlic-free-indian/{c['folder']}/",
            f"{c['folder']}/excel/{c['id']}-recipes.xlsx",
            "Starter, Main, Side, Bread, Sweet, Dessert, Salad",
        ]
        for i, v in enumerate(values, 1):
            cell = ws.cell(row, i, v)
            cell.alignment = WRAP
            cell.border = THIN
            cell.font = BODY_FONT
            cell.fill = CREAM if row % 2 == 0 else WHITE
        ws.row_dimensions[row].height = 36
        row += 1
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=6)
    t = ws.cell(row + 1, 1, f"Total recipes: {len(recipes)}  ·  Continents: {len(CONTINENTS)}  ·  Rebuild: python3 scripts/build_continent_recipes.py")
    t.fill = GREEN_SOFT
    t.font = BOLD
    for col in range(1, 7):
        ws.cell(row + 1, col).fill = GREEN_SOFT
    set_widths(ws, {1: 22, 2: 42, 3: 12, 4: 48, 5: 42, 6: 48})
    ws.freeze_panes = "A8"


def build_continent_workbook(continent: dict, recipes: list[dict], path: Path):
    wb = Workbook()
    menu = wb.active
    menu.title = "Menu"
    write_menu_sheet(menu, continent, recipes)
    rules = wb.create_sheet("Rules")
    write_rules_sheet(rules)
    shop = wb.create_sheet("Shopping list")
    write_shopping_sheet(shop, recipes, f"{continent['name']} shopping list")
    table = wb.create_sheet("All dishes")
    rows = [recipe_row(rec, continent) for rec in recipes]
    write_table_sheet(
        table,
        f"{continent['name']} — all dishes",
        continent["community"],
        rows,
    )
    for rec in recipes:
        title = sheet_title(f"{rec['category'][:3]} {rec['name']}")
        ws = wb.create_sheet(title)
        write_recipe_sheet(ws, rec, continent)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def build_master_workbook(recipes: list[dict], path: Path):
    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    write_cover_sheet(cover, recipes)
    rules = wb.create_sheet("Rules")
    write_rules_sheet(rules)
    all_rows = [recipe_row(rec, continent_by_id(rec["continent_id"])) for rec in recipes]
    table = wb.create_sheet("All Recipes")
    write_table_sheet(
        table,
        "All continents — filterable recipe table",
        "Use the drop-downs on row 5 to filter by continent or course",
        all_rows,
    )
    for c in CONTINENTS:
        recs = [x for x in recipes if x["continent_id"] == c["id"]]
        ws = wb.create_sheet(sheet_title(c["name"]))
        write_table_sheet(ws, c["name"], c["community"], [recipe_row(r, c) for r in recs])
    for cat in ["Starter", "Main", "Side", "Bread", "Sweet", "Dessert", "Salad"]:
        recs = [x for x in recipes if x["category"] == cat]
        ws = wb.create_sheet(sheet_title(f"{cat}s" if not cat.endswith("s") else cat))
        write_table_sheet(
            ws,
            f"{cat} — all continents",
            "Same course, seven kitchens",
            [recipe_row(r, continent_by_id(r["continent_id"])) for r in recs],
        )
    shop = wb.create_sheet("Shopping list")
    write_shopping_sheet(shop, recipes, "Master shopping list — all continents")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def build_shopping_workbook(recipes: list[dict], path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "All continents"
    write_shopping_sheet(ws, recipes, "Master shopping list")
    for c in CONTINENTS:
        recs = [x for x in recipes if x["continent_id"] == c["id"]]
        sheet = wb.create_sheet(sheet_title(c["name"]))
        write_shopping_sheet(sheet, recs, f"{c['name']} shopping list")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def continent_readme(continent: dict, recipes: list[dict]) -> str:
    lines = [
        f"# {continent['name']}",
        "",
        f"**Kitchen:** {continent['community']}  ",
        f"**Diet:** vegetarian · no onion · no garlic · no aluminium cookware",
        "",
        continent["note"],
        "",
        f"Excel: [`excel/{continent['id']}-recipes.xlsx`](excel/{continent['id']}-recipes.xlsx)",
        "",
        "| Course | Recipe | File |",
        "|--------|--------|------|",
    ]
    order = ["Starter", "Main", "Side", "Bread", "Sweet", "Dessert", "Salad"]
    for cat in order:
        for rec in recipes:
            if rec["category"] == cat:
                rel = f"{CAT_FOLDERS[cat]}/{slug(rec['name'])}.md"
                lines.append(f"| {cat} | {rec['name']} | [{rel}]({rel}) |")
    lines.append("")
    return "\n".join(lines)


def write_markdown_tree(recipes: list[dict]):
    OUT.mkdir(parents=True, exist_ok=True)
    # Only rebuild continent folders. Never delete india-states/ or other collections.
    keep_roots = {"india-states", "focus-states", "excel"}
    for child in list(OUT.iterdir()):
        if child.name in keep_roots:
            continue
        if child.name.endswith(".md"):
            continue
        if child.is_dir() and child.name[:2].isdigit():
            for path in sorted(child.rglob("*"), reverse=True):
                if path.is_file():
                    path.unlink()
                elif path.is_dir():
                    path.rmdir()
            child.rmdir()
    (OUT / "README.md").write_text(README_MD, encoding="utf-8")
    (OUT / "COOKWARE-AND-DIET-RULES.md").write_text(RULES_MD, encoding="utf-8")
    for c in CONTINENTS:
        recs = [x for x in recipes if x["continent_id"] == c["id"]]
        base = OUT / c["folder"]
        (base / "excel").mkdir(parents=True, exist_ok=True)
        (base / "README.md").write_text(continent_readme(c, recs), encoding="utf-8")
        for rec in recs:
            cat_dir = base / CAT_FOLDERS[rec["category"]]
            cat_dir.mkdir(parents=True, exist_ok=True)
            (cat_dir / f"{slug(rec['name'])}.md").write_text(md_for(rec, c), encoding="utf-8")


def main():
    recipes = all_recipes()
    expected = len(CONTINENTS) * 7
    if len(recipes) != expected:
        raise SystemExit(f"Expected {expected} recipes, got {len(recipes)}")
    cats = {c for c in CAT_FOLDERS}
    for rec in recipes:
        if rec["category"] not in cats:
            raise SystemExit(f"Bad category: {rec}")
        if rec["continent_id"] not in {c["id"] for c in CONTINENTS}:
            raise SystemExit(f"Bad continent: {rec}")
        for word in ("onion", "garlic", "shallot", "leek", "chive"):
            blob = " ".join(
                [rec["name"], rec["why"], rec["notes"], *rec["ingredients"], *rec["method"]]
            ).lower()
            # allow the words only as "no onion" / "without garlic" prohibitions
            if re.search(rf"\b{word}s?\b", blob):
                if not re.search(
                    rf"(no|without|minus|not add|skip|forbidden|never).{{0,40}}\b{word}",
                    blob,
                ) and not re.search(rf"\b{word}.{{0,40}}(free|removed|forbidden)", blob):
                    # still allow "onion-garlic free" style already handled; flag real ingredients
                    if re.search(rf"\b{word} (powder|paste|chopped|sliced|diced|cloves?)\b", blob):
                        raise SystemExit(f"Allium ingredient in {rec['name']}: {word}")

    write_markdown_tree(recipes)
    excel_dir = OUT / "excel"
    excel_dir.mkdir(parents=True, exist_ok=True)
    master = excel_dir / "ALL-CONTINENTS.xlsx"
    shopping = excel_dir / "SHOPPING-LIST.xlsx"
    build_master_workbook(recipes, master)
    build_shopping_workbook(recipes, shopping)
    for c in CONTINENTS:
        recs = [x for x in recipes if x["continent_id"] == c["id"]]
        build_continent_workbook(
            c, recs, OUT / c["folder"] / "excel" / f"{c['id']}-recipes.xlsx"
        )
    md_count = len(list(OUT.rglob("*.md")))
    xlsx_count = len(list(OUT.rglob("*.xlsx")))
    print(f"Wrote {len(recipes)} recipes")
    print(f"Markdown files: {md_count}")
    print(f"Excel workbooks: {xlsx_count}")
    print(f"Root: {OUT}")


if __name__ == "__main__":
    main()
