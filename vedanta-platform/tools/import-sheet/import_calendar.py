#!/usr/bin/env python3
"""
Dry-run importer for The Vedanta Calendar Google Sheet (.xlsx export).

Reads every '<year> Calendar' sheet and extracts group bookings, plus operational
notes (contractors, team, holidays). Writes JSON + a human report. Never touches
the database: review the report, fix the sheet or the mapping, run again.

Usage: python3 import_calendar.py <sheet.xlsx> <out_dir>
"""
import sys, json, re, os, datetime as dt
from collections import Counter
from openpyxl import load_workbook

# Header → key. Layouts differ by year, so columns are located from the header row of each sheet.
HEADERS = [
    ("month", ("month", "google drive")), ("day", ("date",)), ("weekday", ("day",)), ("slot", ("time",)),
    ("holiday", ("national",)), ("sales", ("notes for sales",)), ("trip", ("trip", "retreat booking")),
    ("interest", ("interest",)), ("team", ("team updates",)), ("notes", ("notes",)),
    ("guests", ("number of guests", "overall number")), ("attendees", ("number of attendee",)),
    ("rooms", ("number of rooms", "number of bedrooms")), ("form_link", ("booking form link",)),
    ("form_status", ("booking form status",)), ("contact", ("main contact",)), ("phone_email", ("phone number",)),
    ("package", ("package",)), ("site_visit", ("site visit",)), ("arrival_time", ("arrival time",)), ("leave_time", ("leave time",)),
    ("deposit", ("deposit",)), ("billing", ("billing",)), ("terms", ("signed t&c",)), ("retreat_type", ("type of retreat",)), ("feedback", ("feedback",)),
]
def map_columns(header_row):
    col = {}
    for j, h in enumerate(header_row):
        if not h: continue
        hl = str(h).strip().lower().replace("\n", " ")
        for key, needles in HEADERS:
            if key in col: continue
            if any(hl.startswith(n) or (n in hl and key not in ("weekday", "day", "notes")) for n in needles):
                if key == "weekday" and hl != "day": continue
                if key == "day" and hl != "date": continue
                if key == "notes" and hl != "notes": continue
                col[key] = j; break
    return col
COL = {}
MONTHS = {m: i for i, m in enumerate(
    ["january","february","march","april","may","june","july","august","september","october","november","december"], 1)}

def cell(row, key):
    if key not in COL: return None
    v = row[COL[key]] if COL[key] < len(row) else None
    if v is None: return None
    s = str(v).strip()
    return s or None

def parse_date(year, month, day):
    try: return dt.date(year, MONTHS[month.strip().lower()], int(float(day)))
    except Exception: return None

def form_status(s):
    if not s: return None
    s = s.lower()
    if "complete" in s or "filled" in s: return "COMPLETE"
    if "sent" in s: return "SENT"
    return None

def classify(trip, package, guests):
    t = (trip or "").lower(); p = (package or "").lower()
    if "cancel" in t: return "CANCELLED"
    if "wedding" in t: return "wedding"
    if "day retreat" in t or "day retreat" in p: return "day_retreat"
    if "hire" in p: return "venue_hire"
    if "volunteer" in t: return "volunteer"
    if guests or package: return "residential"
    return None

def split_contact(s):
    if not s: return None, None
    email = re.search(r"[\w.+-]+@[\w-]+\.[\w.-]+", s)
    phone = re.search(r"(\+?\d[\d\s()]{8,}\d)", s)
    return (email.group(0) if email else None), (phone.group(1).strip() if phone else None)

def parse_times(trip):
    """Pull 'Check in: Fri 9th at 5pm' / 'Check out: Sun 11th at 2pm' out of the trip text."""
    out = {}
    if not trip: return out
    for label, key in [("check-?in", "arrival"), ("check-?out", "departure")]:
        m = re.search(label + r"[^\n]*?(\d{1,2}(?::\d{2})?\s*(?:am|pm))", trip, re.I)
        if m: out[key + "_time"] = m.group(1).replace(" ", "").lower()
    return out

def run(path, out_dir):
    wb = load_workbook(path, read_only=True)
    groups, notes, problems = [], [], []
    for ws in wb.worksheets:
        m = re.match(r"(\d{4}) Calendar", ws.title)
        if not m: continue
        year = int(m.group(1)); month = day = None
        global COL
        COL = map_columns(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row = list(row) + [None] * 40
            if "month" in COL and row[COL["month"]]: month = str(row[COL["month"]])
            if "day" in COL and row[COL["day"]] is not None: day = row[COL["day"]]
            slot = (cell(row, "slot") or "").upper().strip()
            slot = slot if slot in ("AM", "PM") else None
            date = parse_date(year, month, day) if month and day is not None else None
            trip, package, guests, rooms = cell(row,"trip"), cell(row,"package"), cell(row,"guests") or cell(row,"attendees"), cell(row,"rooms")
            if guests and len(re.findall(r"\d", str(guests))) > 4: guests = None   # a phone number, not a count
            contact, pe, terms = cell(row,"contact"), cell(row,"phone_email"), cell(row,"terms")
            fstat = form_status(cell(row, "form_status"))
            is_booking = any([contact, package, guests, rooms, fstat]) or (trip and re.search(r"check-?in", trip, re.I))
            if is_booking:
                kind = classify(trip, package, guests)
                email, phone = split_contact(pe or "")
                g = dict(external_ref=f"{year}:{idx}", date=str(date) if date else None, slot=slot,
                         title=(trip or "").split("\n")[0][:120] or None, trip_text=trip,
                         retreat_type=None if kind == "CANCELLED" else kind,
                         status="CANCELLED" if kind == "CANCELLED" else ("CONFIRMED" if fstat == "COMPLETE" else "PROVISIONAL"),
                         expected_guests=guests, expected_rooms=rooms, contact_name=contact,
                         contact_email=email, contact_phone=phone, price_notes=package,
                         spa_access=bool(package and "spa" in package.lower()),
                         use_basis="EXCLUSIVE" if terms and "exclusive" in terms.lower() else ("SHARED" if terms and "shared" in terms.lower() else None),
                         terms_document=terms if terms and ".pdf" in terms.lower() else None,
                         booking_form_status=fstat, deposit=cell(row,"deposit"), billing=cell(row,"billing"),
                         arrival_time_col=cell(row,"arrival_time"), leave_time_col=cell(row,"leave_time"), **parse_times(trip))
                if not g.get("arrival_time") and g["arrival_time_col"]: g["arrival_time"] = str(g["arrival_time_col"])[:10]
                if not g.get("departure_time") and g["leave_time_col"]: g["departure_time"] = str(g["leave_time_col"])[:10]
                if not date: problems.append((g["external_ref"], "no date"))
                if not (contact or trip): problems.append((g["external_ref"], "no contact or title"))
                groups.append(g)
            for key, kind in [("holiday","HOLIDAY"), ("team","TEAM"), ("interest","VIEWING"), ("sales","OTHER")]:
                v = cell(row, key)
                if v and date: notes.append(dict(date=str(date), slot=slot, kind=kind, text=v, external_ref=f"{year}:{idx}:{key}"))
            if trip and not is_booking and date:
                k = "MAINTENANCE" if re.search(r"clean|install|visit|service|hygiene|repair|inspection", trip, re.I) else "OTHER"
                notes.append(dict(date=str(date), slot=slot, kind=k, text=trip, external_ref=f"{year}:{idx}:trip"))
    os.makedirs(out_dir, exist_ok=True)
    json.dump(groups, open(os.path.join(out_dir, "groups.json"), "w"), indent=1)
    json.dump(notes, open(os.path.join(out_dir, "calendar_notes.json"), "w"), indent=1)
    by_year = Counter(g["external_ref"][:4] for g in groups)
    by_type = Counter(g["retreat_type"] or "unclassified" for g in groups)
    orgs = Counter((g["title"] or "").split(" - ")[0][:30] for g in groups if g["title"])
    with open(os.path.join(out_dir, "report.md"), "w") as f:
        f.write("# Sheet import — dry run\n\n")
        f.write(f"Group booking rows found: {len(groups)}  \nOperational notes: {len(notes)}  \nRows needing attention: {len(problems)}\n\n")
        f.write("## By year\n" + "\n".join(f"- {y}: {n}" for y, n in sorted(by_year.items())) + "\n\n")
        f.write("## By type\n" + "\n".join(f"- {t}: {n}" for t, n in by_type.most_common()) + "\n\n")
        f.write("## Most frequent titles\n" + "\n".join(f"- {o}: {n}" for o, n in orgs.most_common(12)) + "\n\n")
        f.write("## Rows needing attention\n" + "\n".join(f"- {r}: {why}" for r, why in problems[:100]) + "\n")
    print(f"groups={len(groups)} notes={len(notes)} problems={len(problems)} -> {out_dir}/report.md")

if __name__ == "__main__":
    run(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else "out")
