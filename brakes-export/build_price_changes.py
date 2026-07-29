#!/usr/bin/env python3
"""Build Brakes_Price_Changes_Since_Account_Start.xlsx from All_Line_Items.csv."""

from __future__ import annotations

import csv
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
SOURCE = ROOT / "All_Line_Items.csv"
OUT_XLSX = ROOT / "Brakes_Price_Changes_Since_Account_Start.xlsx"
OUT_CSV = ROOT / "Price_Changes_Summary.csv"


def load_invoice_lines(path: Path) -> list[dict]:
    rows: list[dict] = []
    with path.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            try:
                price = float(r["Unit Price"])
                qty = float(r["Qty Case"] or 0)
                net = float(r["Net Value"] or 0)
            except ValueError:
                continue
            if (r.get("Is Credit") == "Yes") or price <= 0:
                continue
            rows.append(
                {
                    "date": r["Document Date"],
                    "doc": r["Document Number"],
                    "code": str(r["Brakes Code"]).strip(),
                    "name": (r["Product / Ingredient"] or "").strip(),
                    "qty": qty,
                    "unit_price": price,
                    "net": net,
                    "category": r.get("Category") or "",
                }
            )
    return rows


def parse_date(s: str) -> datetime:
    return datetime.strptime(s, "%Y-%m-%d")


def analyse(inv: list[dict]):
    by_code: dict[str, list] = defaultdict(list)
    for r in inv:
        by_code[r["code"]].append(r)

    price_changes = []
    change_events = []
    price_history = []
    products_with_changes = 0
    total_change_events = 0
    products_single_order = 0

    for code, items in sorted(by_code.items()):
        items_sorted = sorted(items, key=lambda x: (x["date"], x["doc"]))
        name = max((i["name"] for i in items_sorted), key=len) or items_sorted[0]["name"]
        category = items_sorted[-1]["category"] or items_sorted[0]["category"]

        sequence = []
        for i in items_sorted:
            if not sequence or abs(sequence[-1]["price"] - i["unit_price"]) > 0.0001:
                sequence.append(
                    {
                        "date": i["date"],
                        "price": i["unit_price"],
                        "doc": i["doc"],
                        "orders_at_price": 1,
                        "qty_at_price": i["qty"],
                        "spend_at_price": i["net"],
                    }
                )
            else:
                sequence[-1]["orders_at_price"] += 1
                sequence[-1]["qty_at_price"] += i["qty"]
                sequence[-1]["spend_at_price"] += i["net"]

        unique_prices = sorted({round(i["unit_price"], 4) for i in items_sorted})
        first_price = sequence[0]["price"]
        last_price = sequence[-1]["price"]
        n_orders = len(items_sorted)
        total_qty = sum(i["qty"] for i in items_sorted)
        total_spend = sum(i["net"] for i in items_sorted)
        n_changes = len(sequence) - 1
        pct_change = ((last_price - first_price) / first_price * 100) if first_price else 0
        abs_change = last_price - first_price
        avg_price = (total_spend / total_qty) if total_qty else first_price

        if n_orders == 1:
            products_single_order += 1
        if n_changes:
            products_with_changes += 1
            total_change_events += n_changes

        direction = (
            "UP" if abs_change > 0.005 else ("DOWN" if abs_change < -0.005 else "SAME")
        )
        price_changes.append(
            {
                "code": code,
                "name": name,
                "category": category,
                "first_date": items_sorted[0]["date"],
                "last_date": items_sorted[-1]["date"],
                "orders": n_orders,
                "first_price": first_price,
                "last_price": last_price,
                "min_price": min(unique_prices),
                "max_price": max(unique_prices),
                "avg_price": round(avg_price, 4),
                "abs_change": round(abs_change, 4),
                "pct_change": round(pct_change, 2),
                "times_changed": n_changes,
                "distinct_prices": len(unique_prices),
                "total_qty": total_qty,
                "total_spend": round(total_spend, 2),
                "direction": direction,
            }
        )

        for s in sequence:
            price_history.append(
                {
                    "code": code,
                    "name": name,
                    "category": category,
                    "effective_date": s["date"],
                    "unit_price": s["price"],
                    "document": s["doc"],
                    "orders_at_this_price": s["orders_at_price"],
                    "qty_at_this_price": s["qty_at_price"],
                    "spend_at_this_price": round(s["spend_at_price"], 2),
                }
            )

        for i in range(1, len(sequence)):
            prev, cur = sequence[i - 1], sequence[i]
            delta = cur["price"] - prev["price"]
            pct = (delta / prev["price"] * 100) if prev["price"] else 0
            change_events.append(
                {
                    "code": code,
                    "name": name,
                    "category": category,
                    "change_date": cur["date"],
                    "document": cur["doc"],
                    "old_price": prev["price"],
                    "new_price": cur["price"],
                    "abs_change": round(delta, 4),
                    "pct_change": round(pct, 2),
                    "direction": "UP" if delta > 0 else "DOWN",
                    "days_since_prev": (
                        parse_date(cur["date"]) - parse_date(prev["date"])
                    ).days,
                    "prev_price_from": prev["date"],
                }
            )

    price_changes.sort(
        key=lambda x: (-x["times_changed"], -abs(x["pct_change"]), -x["total_spend"])
    )
    change_events.sort(key=lambda x: (x["change_date"], x["code"]))
    return {
        "price_changes": price_changes,
        "change_events": change_events,
        "price_history": price_history,
        "products_analyzed": len(by_code),
        "products_with_changes": products_with_changes,
        "total_change_events": total_change_events,
        "products_single_order": products_single_order,
        "date_min": min(r["date"] for r in inv),
        "date_max": max(r["date"] for r in inv),
        "invoice_lines": len(inv),
    }


def style_header(ws, row: int, cols: int, fill, font, border) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = border


def autosize(ws, min_w=10, max_w=45) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for cell in col:
            if cell.value is not None:
                length = max(length, min(len(str(cell.value)), max_w))
        ws.column_dimensions[letter].width = max(min_w, length + 2)


def write_excel(data) -> None:
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    header_fill = PatternFill("solid", fgColor="1F4E3D")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    title_font = Font(bold=True, name="Calibri", size=16, color="1F4E3D")
    up_fill = PatternFill("solid", fgColor="FCE4D6")
    down_fill = PatternFill("solid", fgColor="E2EFDA")
    money = "£#,##0.00"

    changed_only = [p for p in data["price_changes"] if p["times_changed"] > 0]
    up = sum(1 for p in changed_only if p["direction"] == "UP")
    down = sum(1 for p in changed_only if p["direction"] == "DOWN")
    same_net = sum(1 for p in changed_only if p["direction"] == "SAME")
    stable = data["products_analyzed"] - data["products_with_changes"]

    monthly: dict[str, dict] = defaultdict(lambda: {"events": 0, "up": 0, "down": 0})
    for e in data["change_events"]:
        m = e["change_date"][:7]
        monthly[m]["events"] += 1
        monthly[m]["up" if e["direction"] == "UP" else "down"] += 1

    stats = [
        ("Invoice line items analysed", data["invoice_lines"]),
        ("Unique products ordered", data["products_analyzed"]),
        ("Products ordered only once (no comparison possible)", data["products_single_order"]),
        ("Products with stable price (never changed)", stable),
        ("Products with at least 1 price change", data["products_with_changes"]),
        ("Total price-change events (how many times prices moved)", data["total_change_events"]),
        ("Products ending HIGHER than first price", up),
        ("Products ending LOWER than first price", down),
        ("Products that changed but ended at same as first", same_net),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = "Brakes price changes since account start"
    ws["A1"].font = title_font
    ws["A2"] = "Account: 1832142 — The Vedanta Way Ltd"
    ws["A3"] = "Source: Brakes invoice line items (portal CSV) from All_Line_Items.csv"
    ws["A4"] = f'Period covered: {data["date_min"]} to {data["date_max"]}'
    ws["A5"] = (
        "Method: Unit price compared across invoice dates per Brakes product code. "
        "Consecutive same prices are not counted as a change."
    )
    ws["A7"] = "Answer summary"
    ws["A7"].font = Font(bold=True, size=13, color="1F4E3D")
    ws["A8"] = f'Yes — Brakes prices have changed since invoices began ({data["date_min"]}).'
    ws["A8"].font = Font(bold=True, size=12)
    ws["A10"] = "Metric"
    ws["B10"] = "Value"
    style_header(ws, 10, 2, header_fill, header_font, thin)
    for i, (label, val) in enumerate(stats, start=11):
        ws.cell(row=i, column=1, value=label).border = thin
        ws.cell(row=i, column=2, value=val).border = thin

    ws["A21"] = "Top 15 products by number of price changes"
    ws["A21"].font = Font(bold=True, size=12, color="1F4E3D")
    headers_top = [
        "Brakes Code",
        "Product",
        "Times changed",
        "Distinct prices",
        "First £",
        "Last £",
        "Change £",
        "Change %",
        "Direction",
        "Orders",
        "Spend £",
    ]
    for c, h in enumerate(headers_top, 1):
        ws.cell(row=22, column=c, value=h)
    style_header(ws, 22, len(headers_top), header_fill, header_font, thin)
    for i, p in enumerate(changed_only[:15], start=23):
        vals = [
            p["code"],
            p["name"],
            p["times_changed"],
            p["distinct_prices"],
            p["first_price"],
            p["last_price"],
            p["abs_change"],
            p["pct_change"],
            p["direction"],
            p["orders"],
            p["total_spend"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = thin
            if c in (5, 6, 7, 11):
                cell.number_format = money
            if c == 9:
                cell.fill = up_fill if p["direction"] == "UP" else (
                    down_fill if p["direction"] == "DOWN" else PatternFill()
                )
    autosize(ws)
    ws.column_dimensions["B"].width = 55

    headers = [
        "Brakes Code",
        "Product / Ingredient",
        "Category",
        "First invoice",
        "Last invoice",
        "Orders",
        "First price £",
        "Last price £",
        "Min £",
        "Max £",
        "Avg £",
        "Change £",
        "Change %",
        "Times changed",
        "Distinct prices",
        "Direction",
        "Total spend £",
    ]

    def write_product_sheet(title: str, products: list[dict]) -> None:
        sheet = wb.create_sheet(title)
        for c, h in enumerate(headers, 1):
            sheet.cell(row=1, column=c, value=h)
        style_header(sheet, 1, len(headers), header_fill, header_font, thin)
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(products) + 1}"
        sheet.freeze_panes = "A2"
        for i, p in enumerate(products, start=2):
            vals = [
                p["code"],
                p["name"],
                p["category"],
                p["first_date"],
                p["last_date"],
                p["orders"],
                p["first_price"],
                p["last_price"],
                p["min_price"],
                p["max_price"],
                p["avg_price"],
                p["abs_change"],
                p["pct_change"],
                p["times_changed"],
                p["distinct_prices"],
                p["direction"],
                p["total_spend"],
            ]
            for c, v in enumerate(vals, 1):
                cell = sheet.cell(row=i, column=c, value=v)
                cell.border = thin
                if c in (7, 8, 9, 10, 11, 12, 17):
                    cell.number_format = money
                if c == 16 and title == "Products With Changes":
                    cell.fill = up_fill if p["direction"] == "UP" else (
                        down_fill if p["direction"] == "DOWN" else PatternFill()
                    )
        autosize(sheet)
        sheet.column_dimensions["B"].width = 42

    write_product_sheet("Products With Changes", changed_only)
    write_product_sheet(
        "All Products Prices",
        sorted(data["price_changes"], key=lambda x: (-x["times_changed"], -x["total_spend"])),
    )

    ws3 = wb.create_sheet("Price Change Events")
    headers3 = [
        "Change date",
        "Brakes Code",
        "Product / Ingredient",
        "Category",
        "Old price £",
        "New price £",
        "Change £",
        "Change %",
        "Direction",
        "Days since previous price",
        "Previous price from",
        "Invoice #",
    ]
    for c, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=c, value=h)
    style_header(ws3, 1, len(headers3), header_fill, header_font, thin)
    events = data["change_events"]
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(headers3))}{len(events) + 1}"
    ws3.freeze_panes = "A2"
    for i, e in enumerate(events, start=2):
        vals = [
            e["change_date"],
            e["code"],
            e["name"],
            e["category"],
            e["old_price"],
            e["new_price"],
            e["abs_change"],
            e["pct_change"],
            e["direction"],
            e["days_since_prev"],
            e["prev_price_from"],
            e["document"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=i, column=c, value=v)
            cell.border = thin
            if c in (5, 6, 7):
                cell.number_format = money
            if c == 9:
                cell.fill = up_fill if e["direction"] == "UP" else down_fill
    autosize(ws3)
    ws3.column_dimensions["C"].width = 42

    ws4 = wb.create_sheet("Price History")
    headers4 = [
        "Brakes Code",
        "Product / Ingredient",
        "Category",
        "Effective from (invoice date)",
        "Unit price £",
        "Invoice #",
        "Orders at this price",
        "Qty cases at this price",
        "Spend at this price £",
    ]
    for c, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=c, value=h)
    style_header(ws4, 1, len(headers4), header_fill, header_font, thin)
    hist = data["price_history"]
    ws4.auto_filter.ref = f"A1:{get_column_letter(len(headers4))}{len(hist) + 1}"
    ws4.freeze_panes = "A2"
    for i, h in enumerate(hist, start=2):
        vals = [
            h["code"],
            h["name"],
            h["category"],
            h["effective_date"],
            h["unit_price"],
            h["document"],
            h["orders_at_this_price"],
            h["qty_at_this_price"],
            h["spend_at_this_price"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws4.cell(row=i, column=c, value=v)
            cell.border = thin
            if c in (5, 9):
                cell.number_format = money
    autosize(ws4)
    ws4.column_dimensions["B"].width = 42

    # Move All Products after history for nicer order — recreate month sheet last
    # openpyxl keeps creation order; Overview, Products With Changes, All Products,
    # Events, History is fine. Add Changes By Month.
    ws6 = wb.create_sheet("Changes By Month")
    ws6["A1"] = "Price-change events by month"
    ws6["A1"].font = title_font
    for c, h in enumerate(
        ["Month", "Total change events", "Price increases", "Price decreases"], 1
    ):
        ws6.cell(row=3, column=c, value=h)
    style_header(ws6, 3, 4, header_fill, header_font, thin)
    for i, m in enumerate(sorted(monthly.keys()), start=4):
        d = monthly[m]
        ws6.cell(row=i, column=1, value=m).border = thin
        ws6.cell(row=i, column=2, value=d["events"]).border = thin
        ws6.cell(row=i, column=3, value=d["up"]).border = thin
        ws6.cell(row=i, column=4, value=d["down"]).border = thin
    autosize(ws6)

    # Reorder sheets to preferred layout
    order = [
        "Overview",
        "Products With Changes",
        "Price Change Events",
        "Price History",
        "All Products Prices",
        "Changes By Month",
    ]
    for idx, name in enumerate(order):
        wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))

    wb.save(OUT_XLSX)

    with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Metric", "Value"])
        for label, val in stats:
            w.writerow([label, val])
        w.writerow([])
        w.writerow(
            [
                "Brakes Code",
                "Product",
                "Times changed",
                "First £",
                "Last £",
                "Change %",
                "Direction",
                "Spend £",
            ]
        )
        for p in changed_only:
            w.writerow(
                [
                    p["code"],
                    p["name"],
                    p["times_changed"],
                    p["first_price"],
                    p["last_price"],
                    p["pct_change"],
                    p["direction"],
                    p["total_spend"],
                ]
            )

    print(f"Wrote {OUT_XLSX}")
    print(f"Products with changes: {data['products_with_changes']}")
    print(f"Total change events: {data['total_change_events']}")


def main() -> None:
    inv = load_invoice_lines(SOURCE)
    data = analyse(inv)
    write_excel(data)


if __name__ == "__main__":
    main()
